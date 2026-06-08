# -*- coding: utf-8 -*-
"""
MapFanSim - 全场风机 MAP 仿真工具
主入口：双击编译后的 MapFanSim.exe
开发/打包：Python + PyInstaller
现场运行：无需 Python、无需 .NET、无需 HTA、无需 PowerShell 脚本环境

功能概要：
1. GUI 科技风界面
2. 本地全场 MAP 仿真：本机故障风机 -> 目标正常风机
3. 支持多组仿真关系
4. 支持 rules\风场名\device_maps.csv 多风场规则库
5. 首页选择风场，默认红山嘴风电一场
6. 支持排除 IEMP
6. 支持底部额外项
7. 生成 CSV 对比报告
8. 内置 SFTP 默认连接方式
9. WinSCP / FlashFXP / OMTG / 终端作为备用工具
10. 第三方软件按“目录”配置，启动时自动带工作目录
"""

from __future__ import annotations

import csv
import datetime as _dt
import json
import os
import re
import shutil
import subprocess
import sys
import threading
import time
import traceback
import webbrowser
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

APP_NAME = "MapFanSim"
APP_TITLE = "MapFanSim 全场风机 MAP 仿真工具"
APP_VERSION = "V8-HTA-LikeLayout"
GITHUB_REPOSITORY = "https://github.com/qssec1/map.git"
GITEE_REPOSITORY = "https://gitee.com/qssec/map"
PRODUCT_DOWNLOAD_URL = "https://gitee.com/qssec/map/raw/master/artifacts/MapFanSim-windows-x64.zip"

# -----------------------------
# 路径与基础工具
# -----------------------------

def is_frozen() -> bool:
    return bool(getattr(sys, "frozen", False))


def get_root_dir() -> Path:
    """无论从哪里双击，都以 exe 所在目录作为根目录。"""
    if is_frozen():
        return Path(sys.executable).resolve().parent
    # 源码调试时：src 的上一级作为根目录
    return Path(__file__).resolve().parents[1]

ROOT = get_root_dir()
DIRS = {
    "data": ROOT / "data",
    "input_maps": ROOT / "input_maps",
    "output_maps": ROOT / "output_maps",
    "download": ROOT / "download",
    "update": ROOT / "update",
    "backup": ROOT / "backup",
    "reports": ROOT / "reports",
    "logs": ROOT / "logs",
    "tools": ROOT / "tools",
    "rules": ROOT / "rules",
}

CONFIG_PATH = DIRS["data"] / "config.json"
DEVICE_MAPS_PATH = DIRS["data"] / "device_maps.csv"
RELATIONS_PATH = DIRS["data"] / "relations.csv"
EXTRA_RULES_PATH = DIRS["data"] / "extra_rules.txt"


def ensure_dirs() -> None:
    for p in DIRS.values():
        p.mkdir(parents=True, exist_ok=True)


def now_stamp() -> str:
    return _dt.datetime.now().strftime("%Y%m%d_%H%M%S")


def safe_open_folder(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)
    if sys.platform.startswith("win"):
        os.startfile(str(path))  # type: ignore[attr-defined]
    elif sys.platform == "darwin":
        subprocess.Popen(["open", str(path)])
    else:
        subprocess.Popen(["xdg-open", str(path)])


FARM_SCOPED_DIR_KEYS = ("input_maps", "output_maps", "download", "update", "backup", "reports", "logs")


def _safe_folder_part(value: str, fallback: str = "default") -> str:
    text = str(value or "").strip()
    text = re.sub(r'[\\/:*?"<>|]+', "_", text)
    text = re.sub(r"\s+", "_", text)
    text = text.strip("._ ")
    return text or fallback


def farm_runtime_dir(key: str, farm: Optional[str] = None) -> Path:
    if key not in DIRS:
        raise KeyError(key)
    farm_name = _safe_folder_part(farm or get_current_wind_farm(), DEFAULT_WIND_FARM)
    path = DIRS[key] / farm_name
    path.mkdir(parents=True, exist_ok=True)
    return path


def farm_runtime_path(key: str, *parts: str, farm: Optional[str] = None) -> Path:
    return farm_runtime_dir(key, farm=farm).joinpath(*parts)


def open_farm_runtime_folder(key: str) -> None:
    safe_open_folder(farm_runtime_dir(key))


def ensure_farm_runtime_dirs(farm: Optional[str] = None) -> None:
    for key in FARM_SCOPED_DIR_KEYS:
        farm_runtime_dir(key, farm=farm)


def csv_escape(value: Any) -> str:
    s = "" if value is None else str(value)
    return s


def read_text_lines(path: Path) -> Tuple[List[str], str]:
    """尽量保留中文编码。返回 lines, encoding。"""
    data = path.read_bytes()
    for enc in ("utf-8-sig", "utf-8", "gbk", "gb18030", "latin1"):
        try:
            text = data.decode(enc)
            return text.splitlines(keepends=True), enc
        except UnicodeDecodeError:
            continue
    text = data.decode("latin1", errors="replace")
    return text.splitlines(keepends=True), "latin1"


def write_text_lines(path: Path, lines: List[str], encoding: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    text = "".join(lines)
    try:
        path.write_text(text, encoding=encoding, newline="")
    except Exception:
        path.write_text(text, encoding="utf-8", newline="")


def split_line_keep_newline(line: str) -> Tuple[str, str]:
    if line.endswith("\r\n"):
        return line[:-2], "\r\n"
    if line.endswith("\n"):
        return line[:-1], "\n"
    if line.endswith("\r"):
        return line[:-1], "\r"
    return line, ""


def replace_first_token(line: str, new_value: str) -> Tuple[str, str, str]:
    """替换一行第一个字段，尽量保持原分隔符。返回 new_line, old_first, new_first。"""
    body, nl = split_line_keep_newline(line)
    m = re.match(r"^(\s*)(\S+)(.*)$", body)
    if not m:
        return line, "", new_value
    prefix, old_first, rest = m.group(1), m.group(2), m.group(3)
    return f"{prefix}{new_value}{rest}{nl}", old_first, new_value


def first_token(line: str) -> str:
    body, _ = split_line_keep_newline(line)
    m = re.match(r"^\s*(\S+)", body)
    return m.group(1) if m else ""


def normalize_fan_name(s: str) -> str:
    return s.strip().upper().replace(" ", "")


def parse_int_safe(s: str, default: int = 0) -> int:
    try:
        return int(str(s).strip())
    except Exception:
        return default


# -----------------------------
# 配置结构
# -----------------------------

@dataclass
class Config:
    remoteMode: str = "builtin_sftp"  # builtin_sftp / winscp / flashfxp / terminal / local_only
    host: str = "192.168.149.222"
    port: int = 60022
    username: str = "root"
    password: str = ""
    remoteDir: str = "/opt/goldwind/LEAP2/cfg"
    remoteFile: str = "slaverMB_1.map"
    hostKey: str = ""  # 为空则自动信任

    winscpDir: str = "E:\\WinSCP"
    flashfxpDir: str = "E:\\FlashFXP"
    omtgDir: str = "E:\\OMTG"
    terminalDir: str = ""

    localMapPath: str = ""
    targetMapPath: str = ""
    excludeIemp: bool = True
    mapFileMode: str = "text"  # 预留：text/binary


def load_config() -> Config:
    ensure_dirs()
    if not CONFIG_PATH.exists():
        cfg = Config()
        save_config(cfg)
        return cfg
    try:
        raw = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        base = asdict(Config())
        base.update(raw)
        return Config(**base)
    except Exception:
        # 配置损坏时备份后重建
        bad = CONFIG_PATH.with_suffix(f".bad_{now_stamp()}.json")
        try:
            shutil.copy2(CONFIG_PATH, bad)
        except Exception:
            pass
        cfg = Config()
        save_config(cfg)
        return cfg


def save_config(cfg: Config) -> None:
    ensure_dirs()
    CONFIG_PATH.write_text(json.dumps(asdict(cfg), ensure_ascii=False, indent=2), encoding="utf-8")


# -----------------------------
# 默认数据文件
# -----------------------------

def ensure_default_files() -> None:
    ensure_dirs()
    if not DEVICE_MAPS_PATH.exists():
        with DEVICE_MAPS_PATH.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(["fan", "line_no", "addr", "col2", "col3", "desc", "exclude"])
            # 生成 F1-01FJ 到 F8-20FJ 的占位映射。
            # 真现场使用时，用“规则已内置/查看说明”或手动替换本文件。
            line_no = 1
            for group in range(1, 9):
                for no in range(1, 21):
                    fan = f"F{group}-{no:02d}FJ"
                    # 这里只放少量示例行，真实映射请导入现场 xls/csv。
                    for idx in range(1, 4):
                        w.writerow([fan, line_no, f"ADDR_{fan}_{idx}", f"C2_{idx}", f"C3_{idx}", f"示例点位{idx}", "0"])
                        line_no += 1
    if not RELATIONS_PATH.exists():
        with RELATIONS_PATH.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(["enabled", "local_fan", "target_fan", "note"])
    if not EXTRA_RULES_PATH.exists():
        EXTRA_RULES_PATH.write_text(
            "# 底部额外项，格式：本机行=目标行，例如：900=1280\n"
            "# 也支持范围：900-901=1280-1281\n",
            encoding="utf-8",
        )


# -----------------------------
# 映射与本地处理
# -----------------------------

@dataclass
class MapEntry:
    fan: str
    line_no: int
    addr: str
    col2: str
    col3: str
    desc: str
    exclude: str = "0"

    @property
    def key(self) -> Tuple[str, str, str]:
        return (self.col2.strip(), self.col3.strip(), self.addr.strip())

    def should_exclude(self, exclude_iemp: bool) -> bool:
        ex = str(self.exclude).strip().lower()
        if ex in ("1", "true", "yes", "y", "是"):
            return True
        if exclude_iemp and "IEMP" in self.desc.upper():
            return True
        return False


@dataclass
class Relation:
    enabled: bool
    local_fan: str
    target_fan: str
    note: str = ""


def list_all_fans() -> List[str]:
    fans = []
    for group in range(1, 9):
        for no in range(1, 21):
            fans.append(f"F{group}-{no:02d}FJ")
    return fans


def load_device_maps() -> List[MapEntry]:
    ensure_default_files()
    entries: List[MapEntry] = []
    with DEVICE_MAPS_PATH.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for r in reader:
            entries.append(MapEntry(
                fan=normalize_fan_name(r.get("fan", "")),
                line_no=parse_int_safe(r.get("line_no", "0")),
                addr=str(r.get("addr", "")).strip(),
                col2=str(r.get("col2", "")).strip(),
                col3=str(r.get("col3", "")).strip(),
                desc=str(r.get("desc", "")).strip(),
                exclude=str(r.get("exclude", "0")).strip(),
            ))
    return entries


def load_relations() -> List[Relation]:
    ensure_default_files()
    rels: List[Relation] = []
    with RELATIONS_PATH.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for r in reader:
            note = str(r.get("note", "")).strip()
            enabled = str(r.get("enabled", "1")).strip().lower() not in ("0", "false", "no", "否")
            if is_default_sample_relation(note):
                enabled = False
            rels.append(Relation(
                enabled=enabled,
                local_fan=normalize_fan_name(r.get("local_fan", "")),
                target_fan=normalize_fan_name(r.get("target_fan", "")),
                note=note,
            ))
    return rels


def save_relations(rels: List[Relation]) -> None:
    with RELATIONS_PATH.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["enabled", "local_fan", "target_fan", "note"])
        for r in rels:
            w.writerow(["1" if r.enabled else "0", r.local_fan, r.target_fan, r.note])


def parse_extra_rules(text: str) -> List[Tuple[int, int, str]]:
    """解析 900=1280 或 900-901=1280-1281。返回 local_line, target_line, desc。行号为 1 起始。"""
    rules: List[Tuple[int, int, str]] = []
    for raw in text.splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        if "=" not in line:
            continue
        left, right = [x.strip() for x in line.split("=", 1)]
        m1 = re.match(r"^(\d+)\s*-\s*(\d+)$", left)
        m2 = re.match(r"^(\d+)\s*-\s*(\d+)$", right)
        if m1 and m2:
            a1, a2 = int(m1.group(1)), int(m1.group(2))
            b1, b2 = int(m2.group(1)), int(m2.group(2))
            count = min(abs(a2 - a1), abs(b2 - b1)) + 1
            step_a = 1 if a2 >= a1 else -1
            step_b = 1 if b2 >= b1 else -1
            for i in range(count):
                rules.append((a1 + i * step_a, b1 + i * step_b, raw))
        else:
            try:
                rules.append((int(left), int(right), raw))
            except Exception:
                pass
    return rules


def load_extra_rules_text() -> str:
    ensure_default_files()
    return EXTRA_RULES_PATH.read_text(encoding="utf-8", errors="ignore")


def save_extra_rules_text(text: str) -> None:
    EXTRA_RULES_PATH.write_text(text, encoding="utf-8")


def backup_file(src: Path, reason: str) -> Path:
    backup_dir = farm_runtime_dir("backup")
    dst = backup_dir / f"{src.stem}_{reason}_{now_stamp()}{src.suffix}"
    shutil.copy2(src, dst)
    return dst


def backup_original_name(src: Path) -> Path:
    dst = farm_runtime_path("backup", src.name)
    shutil.copy2(src, dst)
    return dst


def _fan_backup_part(name: str) -> str:
    nums = re.findall(r"\d+", str(name))
    if nums:
        return str(int(nums[-1]))
    part = re.sub(r"[^0-9A-Za-z_-]+", "", str(name).strip())
    return part or "fan"


def remote_backup_stem(relations: List["Relation"]) -> str:
    enabled = [r for r in relations if r.enabled and r.local_fan and r.target_fan]
    if not enabled:
        return "manual-before"
    pairs = [f"{_fan_backup_part(r.local_fan)}-{_fan_backup_part(r.target_fan)}" for r in enabled]
    return f"{'_'.join(pairs)}-before"


def is_default_sample_relation(note: str) -> bool:
    text = str(note or "").strip()
    return text.startswith("示例") or text.startswith("默认示例")


def import_mapping_table(src_path: Path) -> int:
    """导入 CSV/XLSX/XLS 映射表到 device_maps.csv。要求至少能识别 fan/line_no/addr/col2/col3/desc/exclude。"""
    ext = src_path.suffix.lower()
    rows: List[Dict[str, Any]] = []

    def normalize_header(h: str) -> str:
        h = str(h or "").strip().lower()
        aliases = {
            "风机": "fan", "风机号": "fan", "fan": "fan", "device": "fan",
            "行": "line_no", "行号": "line_no", "line": "line_no", "line_no": "line_no", "row": "line_no",
            "地址": "addr", "addr": "addr", "address": "addr",
            "b列": "col2", "col2": "col2", "字段2": "col2", "key2": "col2",
            "c列": "col3", "col3": "col3", "字段3": "col3", "key3": "col3",
            "说明": "desc", "描述": "desc", "desc": "desc", "description": "desc",
            "排除": "exclude", "exclude": "exclude",
        }
        return aliases.get(h, h)

    if ext == ".csv":
        with src_path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for r in reader:
                nr = {normalize_header(k): v for k, v in r.items()}
                rows.append(nr)
    elif ext in (".xlsx", ".xlsm"):
        try:
            import openpyxl  # type: ignore
        except Exception as e:
            raise RuntimeError("导入 xlsx 需要 openpyxl。请在开发机执行 pip install openpyxl 后重新打包。") from e
        wb = openpyxl.load_workbook(src_path, read_only=True, data_only=True)
        ws = wb.active
        data = list(ws.iter_rows(values_only=True))
        if not data:
            raise RuntimeError("映射表为空")
        headers = [normalize_header(str(x or "")) for x in data[0]]
        for row in data[1:]:
            rows.append({headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))})
    elif ext == ".xls":
        try:
            import xlrd  # type: ignore
        except Exception as e:
            raise RuntimeError("导入 xls 需要 xlrd。请在开发机执行 pip install xlrd==1.2.0 后重新打包。") from e
        book = xlrd.open_workbook(str(src_path))
        sheet = book.sheet_by_index(0)
        headers = [normalize_header(str(sheet.cell_value(0, c) or "")) for c in range(sheet.ncols)]
        for r in range(1, sheet.nrows):
            rows.append({headers[c]: sheet.cell_value(r, c) for c in range(sheet.ncols)})
    else:
        raise RuntimeError("只支持 csv / xlsx / xls 映射表")

    out_rows = []
    for r in rows:
        fan = normalize_fan_name(str(r.get("fan", "")))
        line_no = parse_int_safe(str(r.get("line_no", "0")))
        addr = str(r.get("addr", "")).strip()
        col2 = str(r.get("col2", "")).strip()
        col3 = str(r.get("col3", "")).strip()
        desc = str(r.get("desc", "")).strip()
        exclude = str(r.get("exclude", "0")).strip()
        if not fan or not line_no:
            continue
        out_rows.append([fan, line_no, addr, col2, col3, desc, exclude])

    if not out_rows:
        raise RuntimeError("没有识别到有效映射行。请确认表头包含 fan/line_no/addr/col2/col3/desc/exclude 或中文别名。")

    backup_file(DEVICE_MAPS_PATH, "device_maps_before_import") if DEVICE_MAPS_PATH.exists() else None
    with DEVICE_MAPS_PATH.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["fan", "line_no", "addr", "col2", "col3", "desc", "exclude"])
        w.writerows(out_rows)
    return len(out_rows)


@dataclass
class ReplaceReportRow:
    relation: str
    local_fan: str
    target_fan: str
    local_line: int
    target_line: int
    addr: str
    col2: str
    col3: str
    old_value: str
    new_value: str
    changed: str
    desc: str
    source: str


def run_local_simulation(
    cfg: Config,
    relations: List[Relation],
    local_map: Path,
    target_map: Optional[Path],
    extra_rules_text: str,
    log_func,
) -> Tuple[Path, Path, int]:
    """执行本地仿真。返回 output_map, report_csv, changes_count。"""
    if not local_map.exists():
        raise RuntimeError(f"本机故障 MAP 不存在：{local_map}")
    if target_map and not target_map.exists():
        raise RuntimeError(f"目标正常 MAP 不存在：{target_map}")

    backup = backup_file(local_map, "local_before_sim")
    log_func(f"已备份本机 MAP：{backup}")

    local_lines, enc = read_text_lines(local_map)
    if target_map:
        target_lines, _ = read_text_lines(target_map)
    else:
        target_lines = local_lines[:]

    entries = load_device_maps()
    by_fan: Dict[str, List[MapEntry]] = {}
    for e in entries:
        by_fan.setdefault(e.fan, []).append(e)

    report_rows: List[ReplaceReportRow] = []
    changes = 0

    enabled_rels = [r for r in relations if r.enabled and r.local_fan and r.target_fan]
    if not enabled_rels:
        raise RuntimeError("没有启用的仿真关系，请先添加 本机故障风机 -> 目标正常风机")

    for rel in enabled_rels:
        local_entries = [e for e in by_fan.get(rel.local_fan, []) if not e.should_exclude(cfg.excludeIemp)]
        target_entries = [e for e in by_fan.get(rel.target_fan, []) if not e.should_exclude(cfg.excludeIemp)]
        target_index: Dict[Tuple[str, str, str], MapEntry] = {e.key: e for e in target_entries}
        log_func(f"处理关系：{rel.local_fan} 仿 {rel.target_fan}，本机映射 {len(local_entries)} 行，目标映射 {len(target_entries)} 行")

        for le in local_entries:
            te = target_index.get(le.key)
            if not te:
                continue
            li = le.line_no - 1
            ti = te.line_no - 1
            if li < 0 or li >= len(local_lines) or ti < 0 or ti >= len(target_lines):
                continue
            oldv = first_token(local_lines[li])
            newv = first_token(target_lines[ti])
            new_line, old_first, new_first = replace_first_token(local_lines[li], newv)
            local_lines[li] = new_line
            changed = "是" if old_first != new_first else "否"
            if changed == "是":
                changes += 1
            report_rows.append(ReplaceReportRow(
                relation=f"{rel.local_fan}->{rel.target_fan}",
                local_fan=rel.local_fan,
                target_fan=rel.target_fan,
                local_line=le.line_no,
                target_line=te.line_no,
                addr=le.addr,
                col2=le.col2,
                col3=le.col3,
                old_value=old_first,
                new_value=new_first,
                changed=changed,
                desc=le.desc,
                source="device_maps.csv",
            ))

    # 底部额外项：直接按行号替换第一个字段
    extra_rules = parse_extra_rules(extra_rules_text)
    for local_line, target_line, raw_desc in extra_rules:
        li = local_line - 1
        ti = target_line - 1
        if li < 0 or li >= len(local_lines) or ti < 0 or ti >= len(target_lines):
            continue
        oldv = first_token(local_lines[li])
        newv = first_token(target_lines[ti])
        new_line, old_first, new_first = replace_first_token(local_lines[li], newv)
        local_lines[li] = new_line
        changed = "是" if old_first != new_first else "否"
        if changed == "是":
            changes += 1
        report_rows.append(ReplaceReportRow(
            relation="extra_rules",
            local_fan="额外项",
            target_fan="额外项",
            local_line=local_line,
            target_line=target_line,
            addr="",
            col2="",
            col3="",
            old_value=old_first,
            new_value=new_first,
            changed=changed,
            desc=raw_desc,
            source="extra_rules.txt",
        ))

    out_name = f"{local_map.stem}_sim_{now_stamp()}{local_map.suffix}"
    output_map = farm_runtime_path("output_maps", out_name)
    write_text_lines(output_map, local_lines, enc)

    # 同步一份到 update，便于直接上传
    update_map = farm_runtime_path("update", cfg.remoteFile)
    shutil.copy2(output_map, update_map)
    report_csv = farm_runtime_path("reports", f"replace_report_{now_stamp()}.csv")
    with report_csv.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["仿真关系", "本机故障风机", "目标正常风机", "本机行号", "目标行号", "地址", "col2", "col3", "旧值", "新值", "是否变化", "说明", "来源"])
        for r in report_rows:
            w.writerow([r.relation, r.local_fan, r.target_fan, r.local_line, r.target_line, r.addr, r.col2, r.col3, r.old_value, r.new_value, r.changed, r.desc, r.source])

    log_func(f"生成仿真 MAP：{output_map}")
    log_func(f"同步待上传 MAP：{update_map}")
    log_func(f"生成替换报告：{report_csv}")
    return output_map, report_csv, changes


# -----------------------------
# 远程连接：内置 SFTP / WinSCP / 工具打开
# -----------------------------

class RemoteClient:
    def __init__(self, cfg: Config, log_func):
        self.cfg = cfg
        self.log = log_func

    def test(self) -> None:
        mode = self.cfg.remoteMode
        if mode == "builtin_sftp":
            self._builtin_sftp_test()
        elif mode == "winscp":
            self._winscp_test()
        elif mode == "flashfxp":
            self.open_flashfxp()
            raise RuntimeError("FlashFXP 模式默认作为人工备用工具。已尝试打开 FlashFXP，请人工确认连接。")
        elif mode == "terminal":
            self.open_ssh_terminal()
            raise RuntimeError("终端模式默认作为人工备用工具。已打开终端，请人工确认连接。")
        elif mode == "local_only":
            raise RuntimeError("当前是仅本地模式，不执行远程连接。")
        else:
            raise RuntimeError(f"未知远程连接方式：{mode}")

    def download(self) -> Path:
        mode = self.cfg.remoteMode
        if mode == "builtin_sftp":
            return self._builtin_sftp_download()
        if mode == "winscp":
            return self._winscp_download()
        raise RuntimeError(f"当前远程方式 {mode} 不支持自动下载。请切换为 内置 SFTP 或 WinSCP。")

    def backup_remote(self, backup_stem: str) -> str:
        mode = self.cfg.remoteMode
        if mode == "builtin_sftp":
            return self._builtin_sftp_backup_remote(backup_stem)
        if mode == "winscp":
            return self._winscp_backup_remote(backup_stem)
        raise RuntimeError(f"当前远程方式 {mode} 不支持自动服务器备份。请切换为 内置 SFTP 或 WinSCP。")

    def upload(self, local_file: Path) -> None:
        mode = self.cfg.remoteMode
        if mode == "builtin_sftp":
            self._builtin_sftp_upload(local_file)
            return
        if mode == "winscp":
            self._winscp_upload(local_file)
            return
        raise RuntimeError(f"当前远程方式 {mode} 不支持自动上传。请切换为 内置 SFTP 或 WinSCP。")

    def delete_remote_backup(self, remote_backup_file: str) -> None:
        remote_backup_file = str(remote_backup_file or "").strip()
        if not remote_backup_file:
            return
        backup_name = self._validate_remote_backup_delete_name(remote_backup_file)
        mode = self.cfg.remoteMode
        if mode == "builtin_sftp":
            self._builtin_sftp_delete_backup(backup_name)
            return
        if mode == "winscp":
            self._winscp_delete_backup(backup_name)
            return
        raise RuntimeError(f"当前远程方式 {mode} 不支持自动删除服务器备份。请切换为 内置 SFTP 或 WinSCP。")

    def _validate_remote_backup_delete_name(self, remote_backup_file: str) -> str:
        normalized = remote_backup_file.replace("\\", "/").strip()
        remote_dir = self.cfg.remoteDir.rstrip("/") + "/"
        if "/" in normalized and not normalized.startswith(remote_dir):
            raise RuntimeError(f"拒绝删除非当前远程目录下的文件：{remote_backup_file}")

        backup_name = Path(normalized).name
        suffix = re.escape(Path(self.cfg.remoteFile).suffix or ".map")
        backup_pattern = rf"\d+-\d+(?:_\d+-\d+)*-before{suffix}"
        if (
            not backup_name
            or backup_name == self.cfg.remoteFile
            or not re.fullmatch(backup_pattern, backup_name)
        ):
            raise RuntimeError(f"拒绝删除非本程序服务器备份文件：{remote_backup_file}")
        return backup_name

    def _paramiko_connect(self):
        try:
            import paramiko  # type: ignore
        except Exception as e:
            raise RuntimeError("内置 SFTP 需要 paramiko。请在开发机安装 requirements.txt 后重新打包。") from e

        client = paramiko.SSHClient()
        if not self.cfg.hostKey.strip():
            client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        else:
            # 有 hostKey 时仍先允许连接，连接后比对指纹字符串。
            client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        self.log(f"内置 SFTP 连接 {self.cfg.host}:{self.cfg.port} ...")
        client.connect(
            hostname=self.cfg.host,
            port=int(self.cfg.port),
            username=self.cfg.username,
            password=self.cfg.password,
            timeout=12,
            banner_timeout=12,
            auth_timeout=12,
            look_for_keys=False,
            allow_agent=False,
        )
        # 简单 HostKey 比对：支持用户粘贴指纹片段，空则信任。
        hk = self.cfg.hostKey.strip()
        if hk:
            key = client.get_transport().get_remote_server_key()  # type: ignore[union-attr]
            fp = key.get_fingerprint().hex(":")
            if hk.lower().replace(" ", "") not in fp.lower().replace(" ", ""):
                client.close()
                raise RuntimeError(f"Host Key 不匹配。服务器指纹：{fp}")
        return client

    def _builtin_sftp_test(self) -> None:
        client = self._paramiko_connect()
        sftp = client.open_sftp()
        sftp.chdir(self.cfg.remoteDir)
        sftp.stat(self.cfg.remoteFile)
        sftp.close()
        client.close()
        self.log("内置 SFTP 测试成功。")

    def _builtin_sftp_download(self) -> Path:
        client = self._paramiko_connect()
        sftp = client.open_sftp()
        remote_path = self.cfg.remoteDir.rstrip("/") + "/" + self.cfg.remoteFile
        local_path = farm_runtime_path("download", self.cfg.remoteFile)
        self.log(f"下载：{remote_path} -> {local_path}")
        sftp.get(remote_path, str(local_path))
        sftp.close()
        client.close()
        self.log(f"下载完成：{local_path}")
        return local_path

    def _builtin_sftp_upload(self, local_file: Path) -> None:
        if not local_file.exists():
            raise RuntimeError(f"待上传文件不存在：{local_file}")
        client = self._paramiko_connect()
        sftp = client.open_sftp()
        remote_path = self.cfg.remoteDir.rstrip("/") + "/" + self.cfg.remoteFile
        self.log(f"上传：{local_file} -> {remote_path}")
        sftp.put(str(local_file), remote_path)
        sftp.close()
        client.close()
        self.log("上传完成。")

    def _builtin_sftp_delete_backup(self, backup_name: str) -> None:
        client = self._paramiko_connect()
        sftp = client.open_sftp()
        remote_path = self.cfg.remoteDir.rstrip("/") + "/" + backup_name
        self.log(f"删除服务器备份：{remote_path}")
        try:
            sftp.remove(remote_path)
        finally:
            sftp.close()
            client.close()
        self.log("服务器备份已删除。")

    # ---- WinSCP 备用 ----
    def _builtin_sftp_backup_remote(self, backup_stem: str) -> str:
        client = self._paramiko_connect()
        sftp = client.open_sftp()
        remote_dir = self.cfg.remoteDir.rstrip("/")
        remote_path = remote_dir + "/" + self.cfg.remoteFile
        suffix = Path(self.cfg.remoteFile).suffix
        backup_name = re.sub(r"[^0-9A-Za-z_.-]+", "_", backup_stem) + suffix
        backup_path = remote_dir + "/" + backup_name
        self.log(f"服务器备份：{remote_path} -> {backup_path}")
        with sftp.open(remote_path, "rb") as src, sftp.open(backup_path, "wb") as dst:
            while True:
                chunk = src.read(1024 * 1024)
                if not chunk:
                    break
                dst.write(chunk)
        sftp.close()
        client.close()
        self.log(f"服务器备份完成：{backup_path}")
        return backup_path

    def _find_in_dir(self, base_dir: str, names: List[str]) -> Optional[Path]:
        if not base_dir:
            return None
        base = Path(base_dir)
        for name in names:
            p = base / name
            if p.exists():
                return p
        return None

    def _find_winscp_com(self) -> Path:
        candidates_dirs = [
            self.cfg.winscpDir,
            str(DIRS["tools"] / "WinSCP"),
            str(DIRS["tools"]),
            r"E:\WinSCP",
            r"D:\WinSCP",
            r"C:\Program Files (x86)\WinSCP",
            r"C:\Program Files\WinSCP",
        ]
        for d in candidates_dirs:
            p = self._find_in_dir(d, ["WinSCP.com", "winscp.com"])
            if p:
                return p
        raise RuntimeError("未找到 WinSCP.com。请在设置页填写 WinSCP 所在目录，例如 E:\\WinSCP。注意只填目录。")

    def _run_winscp_script(self, commands: List[str]) -> None:
        exe = self._find_winscp_com()
        script_path = farm_runtime_path("logs", f"winscp_script_{now_stamp()}.txt")
        log_path = farm_runtime_path("logs", f"winscp_{now_stamp()}.log")
        script_text = "\n".join([self._winscp_open_cmd(), "option batch abort", "option confirm off"] + commands + ["exit", ""]) 
        script_path.write_text(script_text, encoding="utf-8")
        cmd = [str(exe), f"/script={script_path}", f"/log={log_path}", "/ini=nul"]
        self.log(f"调用 WinSCP：{exe}")
        proc = subprocess.run(cmd, cwd=str(exe.parent), capture_output=True, text=True, errors="replace", timeout=120)
        if proc.returncode != 0:
            self.log(proc.stdout)
            self.log(proc.stderr)
            raise RuntimeError(f"WinSCP 执行失败，退出码 {proc.returncode}，日志：{log_path}")
        self.log(f"WinSCP 执行成功，日志：{log_path}")

    def _winscp_open_cmd(self) -> str:
        user = self.cfg.username.replace("@", "%40")
        # 密码可能有特殊字符，WinSCP 脚本里用 -password 更稳
        return f'open sftp://{user}@{self.cfg.host}:{self.cfg.port}/ -password="{self.cfg.password}" -hostkey=*'

    def _winscp_test(self) -> None:
        cmds = [f'cd "{self.cfg.remoteDir}"', f'ls "{self.cfg.remoteFile}"']
        self._run_winscp_script(cmds)
        self.log("WinSCP 测试成功。")

    def _winscp_download(self) -> Path:
        local_path = farm_runtime_path("download", self.cfg.remoteFile)
        cmds = [f'cd "{self.cfg.remoteDir}"', f'get "{self.cfg.remoteFile}" "{local_path}"']
        self._run_winscp_script(cmds)
        return local_path

    def _winscp_backup_remote(self, backup_stem: str) -> str:
        suffix = Path(self.cfg.remoteFile).suffix
        backup_name = re.sub(r"[^0-9A-Za-z_.-]+", "_", backup_stem) + suffix
        local_backup = farm_runtime_path("logs", self.cfg.remoteFile)
        cmds = [
            f'cd "{self.cfg.remoteDir}"',
            f'get "{self.cfg.remoteFile}" "{local_backup}"',
            f'put "{local_backup}" "{backup_name}"',
        ]
        self._run_winscp_script(cmds)
        remote_path = self.cfg.remoteDir.rstrip("/") + "/" + backup_name
        self.log(f"服务器备份完成：{remote_path}")
        return remote_path

    def _winscp_upload(self, local_file: Path) -> None:
        remote_path = self.cfg.remoteDir.rstrip("/") + "/" + self.cfg.remoteFile
        cmds = [
            f'cd "{self.cfg.remoteDir}"',
            f'put -nopermissions -nopreservetime -resumesupport=off "{local_file}" "{remote_path}"',
            f'ls "{self.cfg.remoteFile}"',
        ]
        self._run_winscp_script(cmds)

    def _winscp_delete_backup(self, backup_name: str) -> None:
        cmds = [
            f'cd "{self.cfg.remoteDir}"',
            f'rm "{backup_name}"',
        ]
        self.log(f"删除服务器备份：{self.cfg.remoteDir.rstrip('/')}/{backup_name}")
        self._run_winscp_script(cmds)
        self.log("服务器备份已删除。")

    # ---- 第三方工具打开：全部带工作目录 ----
    def _open_exe_from_dir(self, base_dir: str, names: List[str], tool_name: str) -> None:
        p = self._find_in_dir(base_dir, names)
        if not p:
            # 常见 tools 兜底
            p = self._find_in_dir(str(DIRS["tools"] / tool_name), names) or self._find_in_dir(str(DIRS["tools"]), names)
        if not p:
            raise RuntimeError(f"未找到 {tool_name} 主程序。请在设置页填写 {tool_name} 所在目录。注意只填目录，不填 exe。")
        self.log(f"打开 {tool_name}：{p}，工作目录：{p.parent}")
        subprocess.Popen([str(p)], cwd=str(p.parent), shell=False)

    def open_winscp(self) -> None:
        self._open_exe_from_dir(self.cfg.winscpDir, ["WinSCP.exe", "winscp.exe"], "WinSCP")

    def open_flashfxp(self) -> None:
        self._open_exe_from_dir(self.cfg.flashfxpDir, ["FlashFXP.exe", "flashfxp.exe"], "FlashFXP")

    def open_omtg(self) -> None:
        self._open_exe_from_dir(self.cfg.omtgDir, ["OMTG.exe", "OMTGo.exe", "OMTGO.exe", "omtgo.exe"], "OMTG")

    def open_ssh_terminal(self) -> None:
        cmd = f'ssh {self.cfg.username}@{self.cfg.host} -p {self.cfg.port}'
        if sys.platform.startswith("win"):
            subprocess.Popen(["cmd", "/k", cmd], cwd=str(ROOT))
        else:
            subprocess.Popen(cmd, shell=True, cwd=str(ROOT))
        self.log(f"已打开 SSH 终端：{cmd}")

    def open_sftp_terminal(self) -> None:
        cmd = f'sftp -P {self.cfg.port} {self.cfg.username}@{self.cfg.host}'
        if sys.platform.startswith("win"):
            subprocess.Popen(["cmd", "/k", cmd], cwd=str(ROOT))
        else:
            subprocess.Popen(cmd, shell=True, cwd=str(ROOT))
        self.log(f"已打开 SFTP 终端：{cmd}")


# -----------------------------
# GUI
# -----------------------------

class App(tk.Tk):
    def __init__(self):
        ensure_default_files()
        super().__init__()
        self.cfg = load_config()
        self.relation_scope = "local"
        self.local_relations = load_relations_for_scope("local")
        self.cloud_relations = load_relations_for_scope("cloud")
        self.relations = self.local_relations
        self.extra_text_cache = load_extra_rules_text()
        self.relation_trees = []
        self.relation_pick_combos = []
        self.relation_listboxes = []
        self.fan_combos = []
        self.relation_manager = None
        self.relation_manager_listbox = None
        self.relation_count_vars = []
        self.relation_pick_var = tk.StringVar(value="")
        self.log_text_widgets = []
        self.title(f"{APP_TITLE}  {APP_VERSION}")
        self.geometry("1180x760")
        self.minsize(1080, 660)
        self.configure(bg="#0f1720")
        self.protocol("WM_DELETE_WINDOW", self.on_close)

        self.style = ttk.Style(self)
        try:
            self.style.theme_use("clam")
        except Exception:
            pass
        self.setup_style()

        self._build_layout()
        self.refresh_relations_table()
        self.log(f"{APP_TITLE} 已启动")
        self.log(f"根目录：{ROOT}")

    def setup_style(self):
        bg = "#0f1720"
        panel = "#152231"
        panel2 = "#1b2d3f"
        fg = "#e8f3f8"
        accent = "#5fd0ff"
        self.style.configure("TFrame", background=bg)
        self.style.configure("Panel.TFrame", background=panel)
        self.style.configure("Card.TFrame", background=panel2)
        self.style.configure("TLabel", background=bg, foreground=fg, font=("Microsoft YaHei UI", 10))
        self.style.configure("Title.TLabel", background=bg, foreground=accent, font=("Microsoft YaHei UI", 17, "bold"))
        self.style.configure("Sub.TLabel", background=bg, foreground="#9fb5c7", font=("Microsoft YaHei UI", 10))
        self.style.configure("Panel.TLabel", background=panel, foreground=fg, font=("Microsoft YaHei UI", 10))
        self.style.configure("Card.TLabel", background=panel2, foreground=fg, font=("Microsoft YaHei UI", 10))
        self.style.configure("TButton", font=("Microsoft YaHei UI", 10), padding=(10, 6))
        self.style.configure("Accent.TButton", font=("Microsoft YaHei UI", 11, "bold"), padding=(12, 8))
        self.style.configure("TEntry", fieldbackground="#14283a", foreground="#f4fbff", insertcolor="#5fd0ff")
        self.style.configure("TCombobox", fieldbackground="#14283a", foreground="#f4fbff", arrowcolor="#5fd0ff")
        self.style.configure("Treeview", background="#101c28", fieldbackground="#101c28", foreground="#e8f3f8", rowheight=27, font=("Microsoft YaHei UI", 9))
        self.style.configure("Treeview.Heading", background="#203347", foreground="#9fe4ff", font=("Microsoft YaHei UI", 10, "bold"))
        self.style.map("Treeview", background=[("selected", "#2c6f9e")], foreground=[("selected", "#ffffff")])

    def _build_layout(self):
        top = tk.Frame(self, bg="#0a1119", height=62)
        top.pack(side=tk.TOP, fill=tk.X)
        top.pack_propagate(False)
        tk.Label(top, text="MapFanSim", bg="#0a1119", fg="#5fd0ff", font=("Consolas", 22, "bold")).pack(side=tk.LEFT, padx=22)
        tk.Label(top, text="本地处理清晰优先，云端操作独立执行", bg="#0a1119", fg="#9fb5c7", font=("Microsoft YaHei UI", 10)).pack(side=tk.LEFT, padx=8)
        repo_link = tk.Label(top, text="GitHub: qssec1", bg="#0a1119", fg="#6fb8ff", cursor="hand2", font=("Consolas", 9, "underline"))
        repo_link.pack(side=tk.RIGHT, padx=(8, 20))
        repo_link.bind("<Button-1>", lambda _e: webbrowser.open(GITHUB_REPOSITORY))
        tk.Label(top, text="玄度技术支持", bg="#0a1119", fg="#8be0b3", font=("Microsoft YaHei UI", 10)).pack(side=tk.RIGHT, padx=20)

        main = tk.Frame(self, bg="#0f1720")
        main.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        nav = tk.Frame(main, bg="#101c28", width=172)
        nav.pack(side=tk.LEFT, fill=tk.Y)
        nav.pack_propagate(False)
        self.content = tk.Frame(main, bg="#0f1720")
        self.content.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.pages: Dict[str, tk.Frame] = {}
        nav_items = [
            ("local_replace", "本地替换"),
            ("cloud_replace", "云端替换"),
            ("settings", "设置"),
            ("help", "教程"),
        ]
        self.nav_buttons: Dict[str, tk.Button] = {}
        for key, title in nav_items:
            b = tk.Button(nav, text=title, anchor="w", command=lambda k=key: self.show_page(k),
                          bg="#101c28", fg="#dfeaf0", activebackground="#1d3448", activeforeground="#ffffff",
                          relief=tk.FLAT, font=("Microsoft YaHei UI", 11), padx=16, pady=13)
            b.pack(fill=tk.X, padx=10, pady=(8 if key == "home" else 3, 3))
            self.nav_buttons[key] = b

        self._create_local_replace_page()
        self._create_cloud_replace_page()
        self._create_settings_page()
        self._create_help_page()
        self.show_page("local_replace")

    def make_page(self, key: str) -> tk.Frame:
        frame = tk.Frame(self.content, bg="#0f1720")
        self.pages[key] = frame
        return frame

    def show_page(self, key: str):
        if key == "local_replace":
            self.set_relation_scope("local")
        elif key == "cloud_replace":
            self.set_relation_scope("cloud")
        for p in self.pages.values():
            p.pack_forget()
        self.pages[key].pack(fill=tk.BOTH, expand=True)
        for k, b in self.nav_buttons.items():
            b.configure(bg="#24445e" if k == key else "#101c28", fg="#ffffff" if k == key else "#dfeaf0")

    def set_relation_scope(self, scope: str):
        scope = "cloud" if scope == "cloud" else "local"
        self.relation_scope = scope
        self.relations = self.cloud_relations if scope == "cloud" else self.local_relations
        if hasattr(self, "relation_listboxes"):
            self.refresh_relations_table()

    def save_current_relations(self):
        if self.relation_scope == "cloud":
            self.cloud_relations = self.relations
        else:
            self.local_relations = self.relations
        save_relations_for_scope(self.relation_scope, self.relations)

    def card(self, parent, title: str) -> tk.LabelFrame:
        lf = tk.LabelFrame(parent, text=title, bg="#152231", fg="#9fe4ff", font=("Microsoft YaHei UI", 11, "bold"), padx=12, pady=10, bd=1, relief=tk.GROOVE)
        return lf

    def sync_fan_inputs_to_current_farm(self, fans: Optional[List[str]] = None):
        fans = fans if fans is not None else list_all_fans()
        state = "readonly" if fans else "normal"
        for combo in getattr(self, "fan_combos", []):
            try:
                if combo.winfo_exists():
                    combo.configure(values=fans, state=state)
            except Exception:
                pass
        if not fans:
            return
        if hasattr(self, "local_fan_var"):
            lf = normalize_fan_name(self.local_fan_var.get())
            if lf not in fans:
                self.local_fan_var.set(fans[0])
        if hasattr(self, "target_fan_var"):
            tf = normalize_fan_name(self.target_fan_var.get())
            if tf not in fans:
                self.target_fan_var.set(fans[1] if len(fans) > 1 else fans[0])

    def relation_missing_fans(self, lf: str, tf: str) -> List[str]:
        try:
            maps = load_legacy_device_maps()
        except Exception:
            return []
        if not maps:
            return []
        missing = []
        for fan in (lf, tf):
            if fan and fan not in maps and fan not in missing:
                missing.append(fan)
        return missing

    def _create_farm_bar(self, parent):
        farm_card = self.card(parent, "风场")
        farm_card.pack(fill=tk.X, padx=24, pady=(0, 10))
        tk.Label(farm_card, text="当前风场", bg="#152231", fg="#e8f3f8", font=("Microsoft YaHei UI", 10)).pack(side=tk.LEFT, padx=(8, 6), pady=8)
        if not hasattr(self, "wind_farm_var"):
            self.wind_farm_var = tk.StringVar(value=get_current_wind_farm())
        combo = ttk.Combobox(farm_card, textvariable=self.wind_farm_var, values=list_wind_farms(), width=34, state="readonly")
        combo.pack(side=tk.LEFT, padx=6, pady=8)
        self.wind_farm_combo = combo
        ttk.Button(farm_card, text="切换", command=self.switch_wind_farm).pack(side=tk.LEFT, padx=6, pady=8)
        if not hasattr(self, "wind_farm_info_var"):
            self.wind_farm_info_var = tk.StringVar(value=wind_farm_summary())
        ttk.Label(farm_card, textvariable=self.wind_farm_info_var, style="Panel.TLabel").pack(side=tk.LEFT, padx=14, pady=8)

    def _create_relation_panel(self, parent):
        fans = list_all_fans()
        if not hasattr(self, "local_fan_var"):
            self.local_fan_var = tk.StringVar(value=fans[0] if fans else "")
        if not hasattr(self, "target_fan_var"):
            self.target_fan_var = tk.StringVar(value=fans[1] if len(fans) > 1 else (fans[0] if fans else ""))
        c = self.card(parent, "仿真关系")
        c.pack(fill=tk.BOTH, expand=True)
        form = tk.Frame(c, bg="#152231")
        form.pack(fill=tk.X)
        tk.Label(form, text="本机故障风机", bg="#152231", fg="#e8f3f8").grid(row=0, column=0, padx=6, pady=8, sticky="w")
        tk.Label(form, text="目标正常风机", bg="#152231", fg="#e8f3f8").grid(row=1, column=0, padx=6, pady=4, sticky="w")
        self.local_fan_combo = ttk.Combobox(form, textvariable=self.local_fan_var, values=fans, width=22, state="readonly" if fans else "normal")
        self.local_fan_combo.grid(row=0, column=1, padx=6, pady=8, sticky="w")
        self.target_fan_combo = ttk.Combobox(form, textvariable=self.target_fan_var, values=fans, width=22, state="readonly" if fans else "normal")
        self.target_fan_combo.grid(row=1, column=1, padx=6, pady=4, sticky="w")
        self.fan_combos.extend([self.local_fan_combo, self.target_fan_combo])
        self.sync_fan_inputs_to_current_farm(fans)
        ttk.Button(form, text="添加关系", command=self.add_relation).grid(row=0, column=2, padx=(16, 6), pady=8, sticky="ew")
        ttk.Button(form, text="删除当前关系", command=self.delete_current_relation).grid(row=1, column=2, padx=(16, 6), pady=4, sticky="ew")
        pick = ttk.Combobox(form, textvariable=self.relation_pick_var, values=[], width=44, state="readonly")
        pick.bind("<<ComboboxSelected>>", lambda _e: self.apply_relation_pick(), add="+")
        self.relation_pick_combos.append(pick)
        ttk.Button(form, text="管理已添加关系", command=self.open_relation_manager).grid(row=0, column=3, rowspan=2, padx=(10, 6), pady=8, sticky="nsew")
        count_var = tk.StringVar(value="已添加关系：0 条")
        self.relation_count_vars.append(count_var)
        tk.Label(form, textvariable=count_var, bg="#152231", fg="#9fb5c7").grid(row=2, column=0, columnspan=4, sticky="w", padx=6, pady=(6, 0))

        visual_wrap = tk.LabelFrame(form, text="已添加仿真关系摘要", bg="#152231", fg="#9fe4ff", padx=6, pady=6)
        visual_wrap.grid(row=2, column=0, columnspan=4, sticky="ew", padx=6, pady=(8, 4))
        visual = tk.Frame(visual_wrap, bg="#101c28", bd=1, relief=tk.SOLID, height=96)
        visual.pack(fill=tk.BOTH, expand=True)
        visual.pack_propagate(False)
        list_wrap = tk.Frame(visual, bg="#101c28")
        list_wrap.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)
        relation_listbox = tk.Listbox(
            list_wrap,
            height=3,
            bg="#0b1722",
            fg="#e8f3f8",
            selectbackground="#2c6f9e",
            selectforeground="#ffffff",
            activestyle="dotbox",
            exportselection=False,
            font=("Microsoft YaHei UI", 10),
        )
        yscroll = ttk.Scrollbar(list_wrap, orient=tk.VERTICAL, command=relation_listbox.yview)
        xscroll = ttk.Scrollbar(visual, orient=tk.HORIZONTAL, command=relation_listbox.xview)
        relation_listbox.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        relation_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        yscroll.pack(side=tk.RIGHT, fill=tk.Y)
        xscroll.pack(side=tk.BOTTOM, fill=tk.X, padx=4)
        relation_listbox.bind("<<ListboxSelect>>", lambda _e, lb=relation_listbox: self.apply_relation_listbox_pick(lb), add="+")
        self.relation_listboxes.append(relation_listbox)
        visual_wrap.grid_remove()
        form.grid_columnconfigure(1, weight=1)
        form.grid_columnconfigure(3, weight=1)

        cols = ("enabled", "local", "target", "note")
        tree = ttk.Treeview(c, columns=cols, show="headings", height=8, selectmode="extended")
        tree.heading("enabled", text="启用")
        tree.heading("local", text="本机故障风机")
        tree.heading("target", text="目标正常风机")
        tree.heading("note", text="说明")
        tree.column("enabled", width=60, anchor="center")
        tree.column("local", width=140, anchor="center")
        tree.column("target", width=140, anchor="center")
        tree.column("note", width=320, anchor="w")
        tree.pack_forget()
        self.relation_trees.append(tree)
        self.relation_tree = tree
        tree.bind("<FocusIn>", lambda _e, t=tree: self.set_active_relation_tree(t))
        tree.bind("<Button-1>", lambda _e, t=tree: self.set_active_relation_tree(t), add="+")
        tree.bind("<ButtonRelease-1>", lambda e, t=tree: self.select_relation_row(e, t), add="+")
        tree.bind("<<TreeviewSelect>>", lambda _e, t=tree: self.set_active_relation_tree(t), add="+")

    def _grid_buttons(self, parent, buttons, columns=3):
        grid = tk.Frame(parent, bg="#152231")
        grid.pack(fill=tk.X)
        for idx, (text, cmd, accent) in enumerate(buttons):
            ttk.Button(grid, text=text, style="Accent.TButton" if accent else "TButton", command=cmd).grid(
                row=idx // columns, column=idx % columns, sticky="ew", padx=4, pady=4
            )
        for col in range(columns):
            grid.grid_columnconfigure(col, weight=1)
        return grid

    def _create_log_panel(self, parent, title):
        log_card = self.card(parent, title)
        log_card.pack(side=tk.BOTTOM, fill=tk.X, padx=24, pady=(0, 12))
        txt = tk.Text(log_card, height=6, bg="#12263a", fg="#eaf7ff", insertbackground="#5fd0ff", font=("Consolas", 10))
        txt.pack(fill=tk.X, expand=False)
        self.log_text_widgets.append(txt)
        self.log_text = txt
        log_btns = tk.Frame(log_card, bg="#152231")
        log_btns.pack(fill=tk.X, pady=(8, 0))
        ttk.Button(log_btns, text="清空日志", command=self.clear_all_logs).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(log_btns, text="打开 logs", command=lambda: open_farm_runtime_folder("logs")).pack(side=tk.LEFT, padx=6)

    def _create_map_file_panel(self, parent, hint):
        if not hasattr(self, "local_map_var"):
            self.local_map_var = tk.StringVar(value=self.cfg.localMapPath)
        if not hasattr(self, "target_map_var"):
            self.target_map_var = tk.StringVar(value=self.cfg.targetMapPath)
        file_card = self.card(parent, "MAP 文件")
        file_card.pack(fill=tk.X, pady=(0, 6))
        file_grid = tk.Frame(file_card, bg="#152231")
        file_grid.pack(fill=tk.X, expand=False)
        self._path_row(file_grid, "本机故障 MAP", self.local_map_var, self.choose_local_map, 0)
        self._path_row(file_grid, "目标正常 MAP", self.target_map_var, self.choose_target_map, 1)
        file_grid.grid_columnconfigure(1, weight=0)

    def _create_local_replace_page(self):
        p = self.make_page("local_replace")
        ttk.Label(p, text="本地替换", style="Title.TLabel").pack(anchor="w", padx=24, pady=(20, 4))
        ttk.Label(p, text="本地处理不会连接服务器；第三方工具用于现场取文件、看文件、辅助本地替换。", style="Sub.TLabel").pack(anchor="w", padx=24, pady=(0, 14))
        self._create_farm_bar(p)
        self._create_log_panel(p, "运行日志（本地替换）")

        body = tk.Frame(p, bg="#0f1720")
        body.pack(fill=tk.BOTH, expand=True, padx=24, pady=(0, 8))
        left = tk.Frame(body, bg="#0f1720")
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        right = tk.Frame(body, bg="#0f1720", width=440)
        right.pack(side=tk.RIGHT, fill=tk.Y, padx=(10, 0))
        right.pack_propagate(False)

        self._create_map_file_panel(left, "本地替换会处理 input_maps 中的 .map，生成 output_maps / reports，并同步 update。")
        self._create_relation_panel(left)

        c = self.card(right, "本地操作")
        c.pack(fill=tk.X, pady=(0, 8))
        self._grid_buttons(c, [
            ("本地批量替换", self.task_local_sim, True),
            ("input_maps", lambda: open_farm_runtime_folder("input_maps"), False),
            ("output_maps", lambda: open_farm_runtime_folder("output_maps"), False),
            ("reports", lambda: open_farm_runtime_folder("reports"), False),
            ("update", lambda: open_farm_runtime_folder("update"), False),
            ("logs", lambda: open_farm_runtime_folder("logs"), False),
        ])

        tools = self.card(right, "第三方工具")
        tools.pack(fill=tk.X)
        self._grid_buttons(tools, [
            ("FlashFXP", lambda: self.run_tool_action("flashfxp"), False),
            ("WinSCP", lambda: self.run_tool_action("winscp"), False),
            ("OMTG", lambda: self.run_tool_action("omtg"), False),
            ("SSH 终端", lambda: self.run_tool_action("ssh"), False),
            ("SFTP 终端", lambda: self.run_tool_action("sftp"), False),
        ])

    def _create_cloud_replace_page(self):
        p = self.make_page("cloud_replace")
        ttk.Label(p, text="云端替换", style="Title.TLabel").pack(anchor="w", padx=24, pady=(20, 4))
        ttk.Label(p, text="云端页保留和本地一致的仿真选择，执行下载、备份、替换、上传和恢复。", style="Sub.TLabel").pack(anchor="w", padx=24, pady=(0, 14))
        self._create_farm_bar(p)
        self._create_log_panel(p, "运行日志（云端替换）")

        body = tk.Frame(p, bg="#0f1720")
        body.pack(fill=tk.BOTH, expand=True, padx=24, pady=(0, 8))
        left = tk.Frame(body, bg="#0f1720")
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        right = tk.Frame(body, bg="#0f1720", width=440)
        right.pack(side=tk.RIGHT, fill=tk.Y, padx=(10, 0))
        right.pack_propagate(False)

        self._create_map_file_panel(left, "云端下载会写入 download；一键仿真会备份、替换、生成报告并上传 update。")
        self._create_relation_panel(left)

        c = self.card(right, "云端操作")
        c.pack(fill=tk.X, pady=(0, 8))
        self._grid_buttons(c, [
            ("一键仿真", self.task_one_click, True),
            ("测试连接", self.task_test_remote, False),
            ("下载服务器 MAP", self.task_download, False),
            ("上传 update", self.task_upload_latest_update, False),
            ("恢复最近备份", self.task_restore_backup, False),
            ("OMTG", lambda: self.run_tool_action("omtg"), False),
            ("backup", lambda: open_farm_runtime_folder("backup"), False),
        ])

        dirs = self.card(right, "云端目录")
        dirs.pack(fill=tk.X)
        self._grid_buttons(dirs, [
            ("download", lambda: open_farm_runtime_folder("download"), False),
            ("update", lambda: open_farm_runtime_folder("update"), False),
            ("backup", lambda: open_farm_runtime_folder("backup"), False),
            ("reports", lambda: open_farm_runtime_folder("reports"), False),
            ("logs", lambda: open_farm_runtime_folder("logs"), False),
        ])

    def _create_home_page(self):
        p = self.make_page("home")
        ttk.Label(p, text="工作台", style="Title.TLabel").pack(anchor="w", padx=24, pady=(20, 4))
        ttk.Label(p, text="日常操作集中在这里：选风场、选 MAP、配关系，然后本地处理或执行云端同步。", style="Sub.TLabel").pack(anchor="w", padx=24, pady=(0, 14))

        farm_card = self.card(p, "风场")
        farm_card.pack(fill=tk.X, padx=24, pady=(0, 10))
        tk.Label(farm_card, text="当前风场", bg="#152231", fg="#e8f3f8", font=("Microsoft YaHei UI", 10)).pack(side=tk.LEFT, padx=(8, 6), pady=8)
        self.wind_farm_var = tk.StringVar(value=get_current_wind_farm())
        self.wind_farm_combo = ttk.Combobox(farm_card, textvariable=self.wind_farm_var, values=list_wind_farms(), width=34, state="readonly")
        self.wind_farm_combo.pack(side=tk.LEFT, padx=6, pady=8)
        ttk.Button(farm_card, text="切换", command=self.switch_wind_farm).pack(side=tk.LEFT, padx=6, pady=8)
        self.wind_farm_info_var = tk.StringVar(value=wind_farm_summary())
        ttk.Label(farm_card, textvariable=self.wind_farm_info_var, style="Panel.TLabel").pack(side=tk.LEFT, padx=14, pady=8)

        if not hasattr(self, "local_map_var"):
            self.local_map_var = tk.StringVar(value=self.cfg.localMapPath)
        if not hasattr(self, "target_map_var"):
            self.target_map_var = tk.StringVar(value=self.cfg.targetMapPath)

        log_card = self.card(p, "运行日志（本地 / 云端）")
        log_card.pack(side=tk.BOTTOM, fill=tk.X, padx=24, pady=(0, 12))
        self.log_text = tk.Text(log_card, height=6, bg="#12263a", fg="#eaf7ff", insertbackground="#5fd0ff", font=("Consolas", 10))
        self.log_text.pack(fill=tk.X, expand=False)
        log_btns = tk.Frame(log_card, bg="#152231")
        log_btns.pack(fill=tk.X, pady=(8, 0))
        ttk.Button(log_btns, text="清空日志", command=lambda: self.log_text.delete("1.0", tk.END)).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(log_btns, text="打开 logs", command=lambda: open_farm_runtime_folder("logs")).pack(side=tk.LEFT, padx=6)

        body = tk.Frame(p, bg="#0f1720")
        body.pack(fill=tk.BOTH, expand=True, padx=24, pady=(0, 8))

        left = tk.Frame(body, bg="#0f1720")
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        right = tk.Frame(body, bg="#0f1720", width=430)
        right.pack(side=tk.RIGHT, fill=tk.Y, padx=(10, 0))
        right.pack_propagate(False)

        file_card = self.card(left, "1. MAP 文件")
        file_card.pack(fill=tk.X, pady=(0, 10))
        file_grid = tk.Frame(file_card, bg="#152231")
        file_grid.pack(fill=tk.X, expand=True)
        self._path_row(file_grid, "本机故障 MAP", self.local_map_var, self.choose_local_map, 0)
        self._path_row(file_grid, "目标正常 MAP", self.target_map_var, self.choose_target_map, 1)
        tk.Label(file_grid, text="先生成 output_maps 和 reports，确认后再执行云端上传。", bg="#152231", fg="#9fb5c7").grid(row=2, column=1, columnspan=3, sticky="w", pady=(0, 4))

        c1 = self.card(left, "2. 仿真关系")
        c1.pack(fill=tk.BOTH, expand=True)

        fans = list_all_fans()
        form = tk.Frame(c1, bg="#152231")
        form.pack(fill=tk.X)
        tk.Label(form, text="本机故障风机", bg="#152231", fg="#e8f3f8").grid(row=0, column=0, padx=6, pady=8, sticky="w")
        tk.Label(form, text="目标正常风机", bg="#152231", fg="#e8f3f8").grid(row=1, column=0, padx=6, pady=4, sticky="w")
        self.local_fan_var = tk.StringVar(value=fans[0] if fans else "")
        self.target_fan_var = tk.StringVar(value=fans[1] if len(fans) > 1 else (fans[0] if fans else ""))
        self.local_fan_combo = ttk.Combobox(form, textvariable=self.local_fan_var, values=fans, width=22, state="readonly" if fans else "normal")
        self.local_fan_combo.grid(row=0, column=1, padx=6, pady=8, sticky="w")
        self.target_fan_combo = ttk.Combobox(form, textvariable=self.target_fan_var, values=fans, width=22, state="readonly" if fans else "normal")
        self.target_fan_combo.grid(row=1, column=1, padx=6, pady=4, sticky="w")
        self.fan_combos.extend([self.local_fan_combo, self.target_fan_combo])
        self.sync_fan_inputs_to_current_farm(fans)
        ttk.Button(form, text="添加关系", command=self.add_relation).grid(row=0, column=2, padx=(16, 6), pady=8, sticky="ew")
        ttk.Button(form, text="删除选中", command=self.delete_selected_relation).grid(row=1, column=2, padx=(16, 6), pady=4, sticky="ew")
        form.grid_columnconfigure(3, weight=1)

        cols = ("enabled", "local", "target", "note")
        self.relation_tree = ttk.Treeview(c1, columns=cols, show="headings", height=10)
        self.relation_tree.heading("enabled", text="启用")
        self.relation_tree.heading("local", text="本机故障风机")
        self.relation_tree.heading("target", text="目标正常风机")
        self.relation_tree.heading("note", text="说明")
        self.relation_tree.column("enabled", width=60, anchor="center")
        self.relation_tree.column("local", width=140, anchor="center")
        self.relation_tree.column("target", width=140, anchor="center")
        self.relation_tree.column("note", width=360, anchor="w")
        self.relation_tree.pack(fill=tk.BOTH, expand=True, pady=(10, 0))

        c2 = self.card(right, "3. 本地替换")
        c2.pack(fill=tk.X, pady=(0, 8))
        local_grid = tk.Frame(c2, bg="#152231")
        local_grid.pack(fill=tk.X)
        local_buttons = [
            ("本地批量替换", self.task_local_sim),
            ("input_maps", lambda: open_farm_runtime_folder("input_maps")),
            ("output_maps", lambda: open_farm_runtime_folder("output_maps")),
            ("reports", lambda: open_farm_runtime_folder("reports")),
            ("logs", lambda: open_farm_runtime_folder("logs")),
            ("FlashFXP", lambda: self.run_tool_action("flashfxp")),
            ("WinSCP", lambda: self.run_tool_action("winscp")),
            ("OMTG", lambda: self.run_tool_action("omtg")),
            ("SSH 终端", lambda: self.run_tool_action("ssh")),
            ("SFTP 终端", lambda: self.run_tool_action("sftp")),
        ]
        for idx, (text, cmd) in enumerate(local_buttons):
            style = "Accent.TButton" if idx == 0 else "TButton"
            ttk.Button(local_grid, text=text, style=style, command=cmd).grid(row=idx // 3, column=idx % 3, sticky="ew", padx=4, pady=4)
        local_grid.grid_columnconfigure(0, weight=1)
        local_grid.grid_columnconfigure(1, weight=1)
        local_grid.grid_columnconfigure(2, weight=1)

        c3 = self.card(right, "4. 云端功能")
        c3.pack(fill=tk.X, pady=(0, 8))
        cloud_grid = tk.Frame(c3, bg="#152231")
        cloud_grid.pack(fill=tk.X)
        cloud_buttons = [
            ("一键仿真", self.task_one_click),
            ("测试连接", self.task_test_remote),
            ("下载服务器 MAP", self.task_download),
            ("上传 update", self.task_upload_latest_update),
            ("恢复最近备份", self.task_restore_backup),
            ("download", lambda: open_farm_runtime_folder("download")),
            ("update", lambda: open_farm_runtime_folder("update")),
            ("backup", lambda: open_farm_runtime_folder("backup")),
        ]
        for idx, (text, cmd) in enumerate(cloud_buttons):
            style = "Accent.TButton" if idx == 0 else "TButton"
            ttk.Button(cloud_grid, text=text, style=style, command=cmd).grid(row=idx // 3, column=idx % 3, sticky="ew", padx=4, pady=4)
        cloud_grid.grid_columnconfigure(0, weight=1)
        cloud_grid.grid_columnconfigure(1, weight=1)
        cloud_grid.grid_columnconfigure(2, weight=1)

    def _create_local_page(self):
        p = self.make_page("local")
        ttk.Label(p, text="本地 MAP 处理", style="Title.TLabel").pack(anchor="w", padx=24, pady=(22, 4))
        ttk.Label(p, text="本地功能不会连接服务器。可选择本机 MAP 与目标 MAP，生成 output_maps 和 reports。", style="Sub.TLabel").pack(anchor="w", padx=24, pady=(0, 18))
        c = self.card(p, "MAP 文件")
        c.pack(fill=tk.X, padx=24, pady=8)

        self.local_map_var = tk.StringVar(value=self.cfg.localMapPath)
        self.target_map_var = tk.StringVar(value=self.cfg.targetMapPath)
        self._path_row(c, "本机故障 MAP", self.local_map_var, self.choose_local_map, 0)
        self._path_row(c, "目标正常 MAP", self.target_map_var, self.choose_target_map, 1)
        tk.Label(c, text="说明：目标 MAP 可以为空，空时从同一个 MAP 内按风机映射取目标行。", bg="#152231", fg="#9fb5c7").grid(row=2, column=1, columnspan=3, sticky="w", pady=4)

        c2 = self.card(p, "底部额外项 / 特殊行")
        c2.pack(fill=tk.BOTH, expand=True, padx=24, pady=8)
        self.extra_text = tk.Text(c2, height=10, bg="#14283a", fg="#f4fbff", insertbackground="#5fd0ff", font=("Consolas", 10))
        self.extra_text.pack(fill=tk.BOTH, expand=True)
        self.extra_text.insert("1.0", self.extra_text_cache)
        btns = tk.Frame(c2, bg="#152231")
        btns.pack(fill=tk.X, pady=8)
        ttk.Button(btns, text="保存额外项", command=self.save_extra_text).pack(side=tk.LEFT, padx=5)
        ttk.Button(btns, text="本地批量替换", style="Accent.TButton", command=self.task_local_sim).pack(side=tk.LEFT, padx=5)
        ttk.Button(btns, text="打开 input_maps", command=lambda: open_farm_runtime_folder("input_maps")).pack(side=tk.LEFT, padx=5)
        ttk.Button(btns, text="打开 output_maps", command=lambda: open_farm_runtime_folder("output_maps")).pack(side=tk.LEFT, padx=5)

    def _path_row(self, parent, label, var, cmd, row):
        tk.Label(parent, text=label, bg="#152231", fg="#e8f3f8").grid(row=row, column=0, padx=6, pady=4, sticky="w")
        ttk.Entry(parent, textvariable=var, width=42).grid(row=row, column=1, padx=6, pady=4, sticky="w")
        ttk.Button(parent, text="选择", command=cmd).grid(row=row, column=2, padx=6, pady=4)
        parent.grid_columnconfigure(1, weight=0)

    def _create_remote_page(self):
        p = self.make_page("remote")
        ttk.Label(p, text="远程服务器", style="Title.TLabel").pack(anchor="w", padx=24, pady=(22, 4))
        ttk.Label(p, text="默认内置 SFTP。WinSCP 作为备用自动传输；FlashFXP/终端更多用于人工排障。", style="Sub.TLabel").pack(anchor="w", padx=24, pady=(0, 18))
        c = self.card(p, "远程操作")
        c.pack(fill=tk.X, padx=24, pady=8)
        for text, cmd in [
            ("测试连接", self.task_test_remote),
            ("下载服务器 MAP", self.task_download),
            ("上传 update 目录 MAP", self.task_upload_latest_update),
            ("一键仿真：下载→备份→替换→报告→上传", self.task_one_click),
            ("取消仿真 / 上传最近备份恢复", self.task_restore_backup),
        ]:
            ttk.Button(c, text=text, command=cmd).pack(side=tk.LEFT, padx=8, pady=14)
        c2 = self.card(p, "目录快捷入口")
        c2.pack(fill=tk.X, padx=24, pady=8)
        for name, d in [("download", "download"), ("update", "update"), ("backup", "backup"), ("reports", "reports"), ("logs", "logs")]:
            ttk.Button(c2, text=f"打开 {name}", command=lambda x=d: open_farm_runtime_folder(x)).pack(side=tk.LEFT, padx=8, pady=14)

    def _create_tools_page(self):
        p = self.make_page("tools")
        ttk.Label(p, text="第三方工具", style="Title.TLabel").pack(anchor="w", padx=24, pady=(22, 4))
        ttk.Label(p, text="第三方软件只填目录。打开时自动带工作目录，避免 OMTG/OMTGo 找不到配置、DLL、插件。", style="Sub.TLabel").pack(anchor="w", padx=24, pady=(0, 18))
        c = self.card(p, "一键打开第三方工具")
        c.pack(fill=tk.X, padx=24, pady=8)
        ttk.Button(c, text="打开 WinSCP", command=lambda: self.run_tool_action("winscp")).pack(side=tk.LEFT, padx=8, pady=14)
        ttk.Button(c, text="打开 FlashFXP", command=lambda: self.run_tool_action("flashfxp")).pack(side=tk.LEFT, padx=8, pady=14)
        ttk.Button(c, text="打开 OMTG / OMTGo", command=lambda: self.run_tool_action("omtg")).pack(side=tk.LEFT, padx=8, pady=14)
        ttk.Button(c, text="打开 SSH 终端", command=lambda: self.run_tool_action("ssh")).pack(side=tk.LEFT, padx=8, pady=14)
        ttk.Button(c, text="打开 SFTP 终端", command=lambda: self.run_tool_action("sftp")).pack(side=tk.LEFT, padx=8, pady=14)

    def _create_settings_page(self):
        p = self.make_page("settings")
        ttk.Label(p, text="设置", style="Title.TLabel").pack(anchor="w", padx=24, pady=(22, 4))
        ttk.Label(p, text="服务器、工具目录和风场规则统一放在这里，主界面只保留日常操作。", style="Sub.TLabel").pack(anchor="w", padx=24, pady=(0, 14))
        self.var_remoteMode = tk.StringVar(value=self.cfg.remoteMode)
        self.var_host = tk.StringVar(value=self.cfg.host)
        self.var_port = tk.StringVar(value=str(self.cfg.port))
        self.var_username = tk.StringVar(value=self.cfg.username)
        self.var_password = tk.StringVar(value=self.cfg.password)
        self.var_remoteDir = tk.StringVar(value=self.cfg.remoteDir)
        self.var_remoteFile = tk.StringVar(value=self.cfg.remoteFile)
        self.var_hostKey = tk.StringVar(value=self.cfg.hostKey)
        self.var_excludeIemp = tk.BooleanVar(value=self.cfg.excludeIemp)

        settings_body = tk.Frame(p, bg="#0f1720")
        settings_body.pack(fill=tk.BOTH, expand=True, padx=24, pady=(0, 10))

        left = tk.Frame(settings_body, bg="#0f1720")
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        right = tk.Frame(settings_body, bg="#0f1720", width=360)
        right.pack(side=tk.RIGHT, fill=tk.Y, padx=(10, 0))
        right.pack_propagate(False)

        c = self.card(left, "服务器与传输方式")
        c.pack(fill=tk.X, pady=(0, 10))

        row = 0
        self._label_entry(c, "远程连接方式", self.var_remoteMode, row, combo=["builtin_sftp", "winscp", "flashfxp", "terminal", "local_only"]); row += 1
        self._label_entry(c, "服务器地址", self.var_host, row); row += 1
        self._label_entry(c, "端口", self.var_port, row); row += 1
        self._label_entry(c, "用户名", self.var_username, row); row += 1
        self._label_entry(c, "密码 明文显示", self.var_password, row); row += 1
        self._label_entry(c, "服务器目录", self.var_remoteDir, row); row += 1
        self._label_entry(c, "服务器文件", self.var_remoteFile, row); row += 1
        self._label_entry(c, "Host Key 可空", self.var_hostKey, row); row += 1
        tk.Checkbutton(c, text="排除 IEMP 行", variable=self.var_excludeIemp, bg="#152231", fg="#e8f3f8", selectcolor="#14283a", activebackground="#152231", activeforeground="#ffffff").grid(row=row, column=1, sticky="w", padx=8, pady=6); row += 1

        c2 = self.card(left, "备用工具目录")
        c2.pack(fill=tk.X)
        self.var_winscpDir = tk.StringVar(value=self.cfg.winscpDir)
        self.var_flashfxpDir = tk.StringVar(value=self.cfg.flashfxpDir)
        self.var_omtgDir = tk.StringVar(value=self.cfg.omtgDir)
        self.var_terminalDir = tk.StringVar(value=self.cfg.terminalDir)
        self._dir_row(c2, "WinSCP 目录", self.var_winscpDir, 0)
        self._dir_row(c2, "FlashFXP 目录", self.var_flashfxpDir, 1)
        self._dir_row(c2, "OMTG/OMTGo 目录", self.var_omtgDir, 2)
        self._dir_row(c2, "终端工具目录 可空", self.var_terminalDir, 3)

        c3 = self.card(right, "风场规则")
        c3.pack(fill=tk.X, pady=(0, 10))
        ttk.Button(c3, text="导入当前风场规则", command=self.import_mapping_gui).pack(fill=tk.X, pady=5)
        ttk.Button(c3, text="打开 rules 目录", command=lambda: safe_open_folder(DIRS["rules"])).pack(fill=tk.X, pady=5)
        ttk.Button(c3, text="打开当前风场目录", command=lambda: safe_open_folder(get_current_farm_dir())).pack(fill=tk.X, pady=5)
        ttk.Button(c3, text="刷新规则统计", command=self.show_mapping_count).pack(fill=tk.X, pady=5)
        self.mapping_info_var = tk.StringVar(value=wind_farm_summary())
        ttk.Label(c3, textvariable=self.mapping_info_var, style="Panel.TLabel", wraplength=310).pack(anchor="w", pady=(8, 0))

        c4 = self.card(right, "项目链接")
        c4.pack(fill=tk.X, pady=(0, 10))
        ttk.Button(c4, text="下载最新版", command=lambda: webbrowser.open(PRODUCT_DOWNLOAD_URL)).pack(fill=tk.X, pady=(0, 6))
        ttk.Button(c4, text="打开码云仓库", command=lambda: webbrowser.open(GITEE_REPOSITORY)).pack(fill=tk.X, pady=4)
        ttk.Button(c4, text="打开 GitHub 仓库", command=lambda: webbrowser.open(GITHUB_REPOSITORY)).pack(fill=tk.X, pady=4)
        ttk.Label(c4, text=f"码云：{GITEE_REPOSITORY}", style="Panel.TLabel", wraplength=310).pack(anchor="w", pady=(8, 0))
        ttk.Label(c4, text=f"GitHub：{GITHUB_REPOSITORY}", style="Panel.TLabel", wraplength=310).pack(anchor="w", pady=(4, 0))
        ttk.Label(c4, text=f"最新版：{PRODUCT_DOWNLOAD_URL}", style="Panel.TLabel", wraplength=310).pack(anchor="w", pady=(4, 0))

        ttk.Button(right, text="保存设置", style="Accent.TButton", command=self.save_settings).pack(fill=tk.X, pady=(4, 0))

    def _label_entry(self, parent, label, var, row, combo=None):
        tk.Label(parent, text=label, bg="#152231", fg="#e8f3f8").grid(row=row, column=0, padx=8, pady=6, sticky="w")
        if combo:
            w = ttk.Combobox(parent, textvariable=var, values=combo, width=34)
        else:
            w = ttk.Entry(parent, textvariable=var, width=56)
        w.grid(row=row, column=1, padx=8, pady=6, sticky="w")

    def _dir_row(self, parent, label, var, row):
        tk.Label(parent, text=label, bg="#152231", fg="#e8f3f8").grid(row=row, column=0, padx=8, pady=6, sticky="w")
        ttk.Entry(parent, textvariable=var, width=72).grid(row=row, column=1, padx=8, pady=6, sticky="ew")
        ttk.Button(parent, text="选择目录", command=lambda v=var: self.choose_dir(v)).grid(row=row, column=2, padx=8, pady=6)
        parent.grid_columnconfigure(1, weight=1)

    def _create_mapping_page(self):
        p = self.make_page("mapping")
        ttk.Label(p, text="风机映射", style="Title.TLabel").pack(anchor="w", padx=24, pady=(22, 4))
        ttk.Label(p, text="规则按风场分开保存：rules\风场名\device_maps.csv。首页选择风场后，程序使用对应风场规则；默认风场为红山嘴风电一场。", style="Sub.TLabel").pack(anchor="w", padx=24, pady=(0, 18))
        c = self.card(p, "映射表操作")
        c.pack(fill=tk.X, padx=24, pady=8)
        ttk.Button(c, text="给当前风场导入规则", command=self.import_mapping_gui).pack(side=tk.LEFT, padx=8, pady=14)
        ttk.Button(c, text="打开 rules 规则目录", command=lambda: safe_open_folder(DIRS["rules"])).pack(side=tk.LEFT, padx=8, pady=14)
        ttk.Button(c, text="刷新映射统计", command=self.show_mapping_count).pack(side=tk.LEFT, padx=8, pady=14)
        self.mapping_info_var = tk.StringVar(value="")
        ttk.Label(c, textvariable=self.mapping_info_var, style="Panel.TLabel").pack(side=tk.LEFT, padx=16)

        c2 = self.card(p, "字段要求")
        c2.pack(fill=tk.BOTH, expand=True, padx=24, pady=8)
        txt = tk.Text(c2, bg="#14283a", fg="#f4fbff", font=("Consolas", 10), wrap="word")
        txt.pack(fill=tk.BOTH, expand=True)
        txt.insert("1.0", "device_maps.csv 必须包含这些列：\n\nfan,line_no,addr,col2,col3,desc,exclude\n\n说明：\n- fan：风机号，例如 F1-01FJ\n- line_no：MAP 文件行号，1 起始\n- addr：地址或唯一标识\n- col2/col3：用于匹配本机与目标的关键字段\n- desc：说明，包含 IEMP 时可被排除\n- exclude：1 表示强制排除\n\n仿真逻辑：\n程序按 col2 + col3 + addr 建索引，找到本机风机和目标风机的对应行，只替换本机 MAP 对应行的第一个字段，并生成 CSV 对比报告。\n")
        txt.config(state="disabled")

    def _create_logs_page(self):
        p = self.make_page("logs")
        ttk.Label(p, text="运行日志", style="Title.TLabel").pack(anchor="w", padx=24, pady=(22, 4))
        self.log_text = tk.Text(p, bg="#12263a", fg="#eaf7ff", insertbackground="#5fd0ff", font=("Consolas", 10))
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=24, pady=12)
        bottom = tk.Frame(p, bg="#0f1720")
        bottom.pack(fill=tk.X, padx=24, pady=(0, 12))
        ttk.Button(bottom, text="清空界面日志", command=lambda: self.log_text.delete("1.0", tk.END)).pack(side=tk.LEFT, padx=5)
        ttk.Button(bottom, text="打开 logs 目录", command=lambda: open_farm_runtime_folder("logs")).pack(side=tk.LEFT, padx=5)

    def _create_help_page(self):
        p = self.make_page("help")
        ttk.Label(p, text="教程", style="Title.TLabel").pack(anchor="w", padx=24, pady=(22, 4))
        link_bar = tk.Frame(p, bg="#0f1720")
        link_bar.pack(fill=tk.X, padx=24, pady=(0, 8))
        ttk.Button(link_bar, text="最新版下载", command=lambda: webbrowser.open(PRODUCT_DOWNLOAD_URL)).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Label(link_bar, text=PRODUCT_DOWNLOAD_URL, style="Sub.TLabel").pack(side=tk.LEFT, padx=12)
        c = self.card(p, "现场使用")
        c.pack(fill=tk.BOTH, expand=True, padx=24, pady=8)
        txt = tk.Text(c, bg="#14283a", fg="#f4fbff", font=("Microsoft YaHei UI", 11), wrap="word")
        txt.pack(fill=tk.BOTH, expand=True)
        txt.insert("1.0", f"""
一、第一次使用
1. 打开“设置”，确认连接类型为内置 SFTP 或 WinSCP。
2. 服务器文件必须保持为 slaverMB_1.map，服务器只识别这个文件名。
3. 填好服务器地址、端口、用户名、密码、服务器目录和服务器文件。
4. 第三方工具只填写目录，不填写 exe 或 com。
5. 规则导入在“设置”里，只在现场规则变更时使用。

二、本地替换
1. 进入“本地替换”，选择风场。
2. 选择本机故障 MAP 和目标正常 MAP，或者把 .map 放到 input_maps。
3. 添加仿真关系，例如 1 仿 2、3 仿 6。
4. 点“本地批量替换”。
5. 新 MAP 生成在 output_maps，待上传文件固定同步到 update\\slaverMB_1.map，报告在 reports。

三、云端替换
1. “下载服务器 MAP”只下载服务器原始 slaverMB_1.map 到 download，文件名不改、内容不改。
2. “一键仿真”会先在服务器备份，再下载原始 MAP 到本地，再本地备份，再替换生成 update\\slaverMB_1.map，最后上传覆盖服务器 slaverMB_1.map。
3. “上传 update”也会先在服务器备份，然后只上传 update\\slaverMB_1.map 覆盖服务器 slaverMB_1.map。
4. “恢复最近备份”会从本地 backup 取最近备份，并上传覆盖服务器 slaverMB_1.map。

四、备份规则
1. 服务器备份不能叫 slaverMB_1.map，避免被服务误识别。
2. 单条仿真关系的服务器备份名示例：1-2-before.map。
3. 多条仿真关系的服务器备份名示例：1-2_3-6-before.map。
4. 本地 backup 里的原始备份保持服务器原文件名：slaverMB_1.map。
5. download 里的下载文件也保持服务器原文件名：slaverMB_1.map。

五、注意事项
- 服务器生效文件只有 slaverMB_1.map，其他名字只作为备份或人工查看。
- 上传前必须先备份服务器原文件。
- 日志在本地替换和云端替换底部，连接、下载、上传、备份问题都先看日志。
- OMTG 可以在本地替换和云端替换页面直接打开。
- Host Key 可空，空时自动信任服务器；密码会保存在 data/config.json。

六、项目仓库
- 码云仓库：{GITEE_REPOSITORY}
- GitHub 仓库：{GITHUB_REPOSITORY}
- 最新版下载：{PRODUCT_DOWNLOAD_URL}
""")
        txt.config(state="disabled")

    # ---- GUI 辅助 ----
    def log(self, msg: str):
        ensure_dirs()
        line = f"[{_dt.datetime.now().strftime('%H:%M:%S')}] {msg}\n"
        for widget in getattr(self, "log_text_widgets", []):
            try:
                widget.insert(tk.END, line)
                widget.see(tk.END)
            except Exception:
                pass
        try:
            with farm_runtime_path("logs", f"run_{_dt.datetime.now().strftime('%Y%m%d')}.log").open("a", encoding="utf-8") as f:
                f.write(line)
        except Exception:
            pass

    def clear_all_logs(self):
        for widget in getattr(self, "log_text_widgets", []):
            try:
                widget.delete("1.0", tk.END)
            except Exception:
                pass

    def run_bg(self, title: str, func):
        def worker():
            self.log(f"开始：{title}")
            try:
                func()
                self.log(f"完成：{title}")
                self.after(0, lambda: messagebox.showinfo("完成", f"{title} 完成"))
            except Exception as e:
                tb = traceback.format_exc()
                self.log(f"失败：{title}：{e}")
                self.log(tb)
                self.after(0, lambda: messagebox.showerror("失败", f"{title} 失败：\n{e}"))
        threading.Thread(target=worker, daemon=True).start()

    def save_settings(self):
        self.cfg.remoteMode = self.var_remoteMode.get().strip()
        self.cfg.host = self.var_host.get().strip()
        self.cfg.port = parse_int_safe(self.var_port.get(), 22)
        self.cfg.username = self.var_username.get().strip()
        self.cfg.password = self.var_password.get()
        self.cfg.remoteDir = self.var_remoteDir.get().strip()
        self.cfg.remoteFile = self.var_remoteFile.get().strip()
        self.cfg.hostKey = self.var_hostKey.get().strip()
        self.cfg.winscpDir = self.var_winscpDir.get().strip()
        self.cfg.flashfxpDir = self.var_flashfxpDir.get().strip()
        self.cfg.omtgDir = self.var_omtgDir.get().strip()
        self.cfg.terminalDir = self.var_terminalDir.get().strip()
        self.cfg.excludeIemp = bool(self.var_excludeIemp.get())
        self.cfg.localMapPath = self.local_map_var.get().strip() if hasattr(self, "local_map_var") else self.cfg.localMapPath
        self.cfg.targetMapPath = self.target_map_var.get().strip() if hasattr(self, "target_map_var") else self.cfg.targetMapPath
        save_config(self.cfg)
        self.log("设置已保存。")
        messagebox.showinfo("保存成功", "设置已保存。")

    def choose_dir(self, var):
        d = filedialog.askdirectory(initialdir=var.get() or str(ROOT))
        if d:
            var.set(d)

    def choose_local_map(self):
        p = filedialog.askopenfilename(initialdir=str(farm_runtime_dir("input_maps")), title="选择本机故障 MAP", filetypes=[("MAP 文件", "*.map"), ("所有文件", "*.*")])
        if p:
            self.local_map_var.set(p)
            self.cfg.localMapPath = p
            save_config(self.cfg)

    def choose_target_map(self):
        p = filedialog.askopenfilename(initialdir=str(farm_runtime_dir("input_maps")), title="选择目标正常 MAP", filetypes=[("MAP 文件", "*.map"), ("所有文件", "*.*")])
        if p:
            self.target_map_var.set(p)
            self.cfg.targetMapPath = p
            save_config(self.cfg)

    def add_relation(self):
        lf = normalize_fan_name(self.local_fan_var.get())
        tf = normalize_fan_name(self.target_fan_var.get())
        missing = self.relation_missing_fans(lf, tf)
        if missing:
            self.sync_fan_inputs_to_current_farm()
            msg = f"当前风场“{get_current_wind_farm()}”没有风机：{'、'.join(missing)}。请从当前风场下拉列表重新选择。"
            self.log(f"添加仿真关系失败：{msg}")
            messagebox.showwarning("提示", msg)
            return
        if lf == tf:
            messagebox.showwarning("提示", "本机故障风机和目标正常风机不能一样。")
            return
        for rel in self.relations:
            if rel.local_fan == lf and rel.target_fan == tf:
                if not rel.enabled:
                    rel.enabled = True
                    rel.note = "GUI 启用"
                    self.save_current_relations()
                    self.refresh_relations_table()
                    self.log(f"启用已有仿真关系：{lf} -> {tf}")
                else:
                    messagebox.showinfo("提示", "这组仿真关系已经存在，不会重复添加。")
                return
        self.relations.append(Relation(True, lf, tf, "GUI 添加"))
        self.save_current_relations()
        self.refresh_relations_table()
        self.select_relation_index(len(self.relations) - 1)
        self.log(f"添加仿真关系：{lf} -> {tf}")

    def delete_selected_relation(self):
        selected = set()
        list_idx = self.get_relation_listbox_index()
        if list_idx is not None:
            selected.add(list_idx)
        else:
            pick_idx = self.get_relation_pick_index()
            if pick_idx is not None:
                selected.add(pick_idx)

        tree = self.get_active_relation_tree()
        if tree is not None:
            for iid in tree.selection():
                try:
                    selected.add(int(iid))
                except ValueError:
                    continue
        if not selected:
            self.log("删除仿真关系：请先在“已添加关系”里选择要删除的关系。")
            return
        self.delete_relation_indexes(selected)

    def delete_current_relation(self):
        lf = normalize_fan_name(self.local_fan_var.get()) if hasattr(self, "local_fan_var") else ""
        tf = normalize_fan_name(self.target_fan_var.get()) if hasattr(self, "target_fan_var") else ""
        selected = {
            idx for idx, rel in enumerate(self.relations)
            if rel.local_fan == lf and rel.target_fan == tf
        }
        if not selected:
            self.log(f"删除仿真关系：没有找到当前关系 {lf} -> {tf}。")
            return
        self.delete_relation_indexes(selected)

    def delete_relation_indexes(self, selected):
        next_idx = min(selected) if selected else None
        indexes = sorted(selected, reverse=True)
        removed = []
        for idx in indexes:
            if 0 <= idx < len(self.relations):
                removed.append(self.relations.pop(idx))
        self.save_current_relations()
        self.refresh_relations_table()
        if self.relations and next_idx is not None:
            self.select_relation_index(min(next_idx, len(self.relations) - 1))
        if removed:
            names = "，".join(f"{r.local_fan}->{r.target_fan}" for r in reversed(removed))
            self.log(f"已删除仿真关系：{names}")

    def relation_label(self, idx: int, rel: Relation) -> str:
        status = "启用" if rel.enabled else "禁用"
        return f"{idx + 1}. {rel.local_fan} -> {rel.target_fan} [{status}]"

    def relation_pick_values(self) -> List[str]:
        return [self.relation_label(idx, rel) for idx, rel in enumerate(self.relations)]

    def get_relation_pick_index(self) -> Optional[int]:
        value = self.relation_pick_var.get().strip() if hasattr(self, "relation_pick_var") else ""
        if not value:
            return None
        m = re.match(r"^(\d+)\.", value)
        if not m:
            return None
        idx = int(m.group(1)) - 1
        return idx if 0 <= idx < len(self.relations) else None

    def apply_relation_pick(self):
        idx = self.get_relation_pick_index()
        if idx is None:
            return
        self.select_relation_index(idx)

    def get_relation_listbox_index(self) -> Optional[int]:
        for listbox in getattr(self, "relation_listboxes", []):
            selection = listbox.curselection()
            if selection:
                idx = int(selection[0])
                return idx if 0 <= idx < len(self.relations) else None
        return None

    def apply_relation_listbox_pick(self, listbox):
        selection = listbox.curselection()
        if not selection:
            return
        idx = int(selection[0])
        if 0 <= idx < len(self.relations):
            self.select_relation_index(idx)

    def open_relation_manager(self):
        if self.relation_manager is not None and self.relation_manager.winfo_exists():
            self.relation_manager.lift()
            self.relation_manager.focus_force()
            self.refresh_relation_manager()
            return

        win = tk.Toplevel(self)
        win.title("管理已添加仿真关系")
        win.geometry("760x420")
        win.minsize(680, 360)
        win.configure(bg="#0f1720")
        win.transient(self)
        self.relation_manager = win

        tk.Label(
            win,
            text="已添加仿真关系",
            bg="#0f1720",
            fg="#5fd0ff",
            font=("Microsoft YaHei UI", 16, "bold"),
        ).pack(anchor="w", padx=16, pady=(14, 8))

        box_frame = tk.Frame(win, bg="#0f1720")
        box_frame.pack(fill=tk.BOTH, expand=True, padx=16, pady=(0, 10))
        lb = tk.Listbox(
            box_frame,
            bg="#0b1722",
            fg="#e8f3f8",
            selectbackground="#2c6f9e",
            selectforeground="#ffffff",
            activestyle="dotbox",
            exportselection=False,
            font=("Microsoft YaHei UI", 11),
        )
        yscroll = ttk.Scrollbar(box_frame, orient=tk.VERTICAL, command=lb.yview)
        xscroll = ttk.Scrollbar(win, orient=tk.HORIZONTAL, command=lb.xview)
        lb.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        yscroll.pack(side=tk.RIGHT, fill=tk.Y)
        xscroll.pack(fill=tk.X, padx=16, pady=(0, 8))
        lb.bind("<<ListboxSelect>>", lambda _e: self.apply_relation_listbox_pick(lb), add="+")
        self.relation_manager_listbox = lb

        buttons = tk.Frame(win, bg="#0f1720")
        buttons.pack(fill=tk.X, padx=16, pady=(0, 14))
        ttk.Button(buttons, text="删除选中关系", command=self.delete_manager_selected_relation).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(buttons, text="关闭", command=win.destroy).pack(side=tk.RIGHT)

        win.protocol("WM_DELETE_WINDOW", win.destroy)
        self.refresh_relation_manager()

    def refresh_relation_manager(self):
        lb = getattr(self, "relation_manager_listbox", None)
        if lb is None or not lb.winfo_exists():
            return
        current = lb.curselection()
        current_idx = int(current[0]) if current else None
        lb.delete(0, tk.END)
        values = self.relation_pick_values()
        if not values:
            lb.insert(tk.END, "暂无已添加关系")
            return
        for value in values:
            lb.insert(tk.END, value)
        if current_idx is not None:
            idx = min(current_idx, len(values) - 1)
            lb.selection_set(idx)
            lb.activate(idx)
            lb.see(idx)

    def delete_manager_selected_relation(self):
        lb = getattr(self, "relation_manager_listbox", None)
        if lb is None or not lb.winfo_exists():
            return
        selection = lb.curselection()
        if not selection:
            self.log("删除仿真关系：请先在管理窗口里选择一条关系。")
            return
        idx = int(selection[0])
        if 0 <= idx < len(self.relations):
            self.delete_relation_indexes({idx})
            self.refresh_relation_manager()

    def render_relation_visual_lists(self):
        values = [
            f"{idx + 1}. 本机：{rel.local_fan}  ->  目标风机：{rel.target_fan}  [{('启用' if rel.enabled else '禁用')}]"
            for idx, rel in enumerate(self.relations)
        ]
        if not values:
            values = ["暂无已添加关系"]
        for listbox in getattr(self, "relation_listboxes", []):
            current = listbox.curselection()
            current_idx = int(current[0]) if current else None
            listbox.delete(0, tk.END)
            for value in values:
                listbox.insert(tk.END, value)
            if self.relations and current_idx is not None:
                idx = min(current_idx, len(self.relations) - 1)
                listbox.selection_set(idx)
                listbox.activate(idx)
                listbox.see(idx)
        self.refresh_relation_manager()

    def set_active_relation_tree(self, tree):
        self.active_relation_tree = tree

    def select_relation_row(self, event, tree):
        self.set_active_relation_tree(tree)
        row = tree.identify_row(event.y)
        if not row:
            return
        tree.focus(row)
        tree.selection_set(row)

    def select_relation_index(self, idx: int):
        iid = str(idx)
        if hasattr(self, "relation_pick_var") and 0 <= idx < len(self.relations):
            rel = self.relations[idx]
            self.relation_pick_var.set(self.relation_label(idx, rel))
            if hasattr(self, "local_fan_var"):
                self.local_fan_var.set(rel.local_fan)
            if hasattr(self, "target_fan_var"):
                self.target_fan_var.set(rel.target_fan)
        for listbox in getattr(self, "relation_listboxes", []):
            if 0 <= idx < listbox.size() and self.relations:
                listbox.selection_clear(0, tk.END)
                listbox.selection_set(idx)
                listbox.activate(idx)
                listbox.see(idx)
        tree = self.get_active_relation_tree()
        if tree is None:
            return
        if iid in tree.get_children():
            tree.focus(iid)
            tree.selection_set(iid)
            tree.see(iid)

    def get_active_relation_tree(self):
        trees = list(getattr(self, "relation_trees", []))
        if not trees and hasattr(self, "relation_tree"):
            trees = [self.relation_tree]

        visible_trees = [tree for tree in trees if tree.winfo_ismapped()]
        active = getattr(self, "active_relation_tree", None)
        if active in visible_trees:
            return active
        for tree in visible_trees:
            if tree.selection():
                self.active_relation_tree = tree
                return tree
        if len(visible_trees) == 1:
            self.active_relation_tree = visible_trees[0]
            return visible_trees[0]
        return None

    def refresh_relations_table(self):
        trees = getattr(self, "relation_trees", [])
        if not trees and hasattr(self, "relation_tree"):
            trees = [self.relation_tree]
        for var in getattr(self, "relation_count_vars", []):
            var.set(f"已添加关系：{len(self.relations)} 条（点击“管理已添加关系”查看、选择、删除）")
        self.render_relation_visual_lists()
        self.refresh_relation_manager()
        values = self.relation_pick_values()
        for combo in getattr(self, "relation_pick_combos", []):
            combo.configure(values=values)
        current = self.relation_pick_var.get().strip() if hasattr(self, "relation_pick_var") else ""
        if values:
            if current not in values:
                self.relation_pick_var.set(values[0])
        elif hasattr(self, "relation_pick_var"):
            self.relation_pick_var.set("")
        if not trees:
            return
        for tree in trees:
            for i in tree.get_children():
                tree.delete(i)
            for idx, r in enumerate(self.relations):
                tree.insert("", tk.END, iid=str(idx), values=("是" if r.enabled else "否", r.local_fan, r.target_fan, r.note))

    def save_extra_text(self):
        text = self.extra_text.get("1.0", tk.END)
        save_extra_rules_text(text)
        self.extra_text_cache = text
        self.log("额外项已保存。")
        messagebox.showinfo("保存成功", "底部额外项已保存。")

    def get_local_paths_for_sim(self) -> Tuple[Path, Optional[Path]]:
        self.save_settings_no_popup()
        local_s = self.local_map_var.get().strip() if hasattr(self, "local_map_var") else self.cfg.localMapPath
        target_s = self.target_map_var.get().strip() if hasattr(self, "target_map_var") else self.cfg.targetMapPath
        if not local_s:
            # 尝试取 input_maps 里最新 map
            maps = sorted(farm_runtime_dir("input_maps").glob("*.map"), key=lambda p: p.stat().st_mtime, reverse=True)
            if maps:
                local_s = str(maps[0])
        if not local_s:
            raise RuntimeError("请选择本机故障 MAP，或把 .map 文件放到 input_maps。")
        local_map = validate_current_farm_map_path(Path(local_s), "本机故障 MAP")
        target_map = validate_current_farm_map_path(Path(target_s), "目标正常 MAP") if target_s else None
        return local_map, target_map

    def save_settings_no_popup(self):
        try:
            self.cfg.remoteMode = self.var_remoteMode.get().strip()
            self.cfg.host = self.var_host.get().strip()
            self.cfg.port = parse_int_safe(self.var_port.get(), 22)
            self.cfg.username = self.var_username.get().strip()
            self.cfg.password = self.var_password.get()
            self.cfg.remoteDir = self.var_remoteDir.get().strip()
            self.cfg.remoteFile = self.var_remoteFile.get().strip()
            self.cfg.hostKey = self.var_hostKey.get().strip()
            self.cfg.winscpDir = self.var_winscpDir.get().strip()
            self.cfg.flashfxpDir = self.var_flashfxpDir.get().strip()
            self.cfg.omtgDir = self.var_omtgDir.get().strip()
            self.cfg.terminalDir = self.var_terminalDir.get().strip()
            self.cfg.excludeIemp = bool(self.var_excludeIemp.get())
        except Exception:
            pass
        if hasattr(self, "local_map_var"):
            self.cfg.localMapPath = self.local_map_var.get().strip()
        if hasattr(self, "target_map_var"):
            self.cfg.targetMapPath = self.target_map_var.get().strip()
        save_config(self.cfg)

    # ---- 任务 ----
    def task_local_sim(self):
        relations = list(self.local_relations)
        def work():
            self.save_settings_no_popup()
            save_relations_for_scope("local", relations)
            extra = self.extra_text.get("1.0", tk.END) if hasattr(self, "extra_text") else load_extra_rules_text()
            save_extra_rules_text(extra)
            maps = sorted(farm_runtime_dir("input_maps").glob("*.map"))
            target_s = self.target_map_var.get().strip() if hasattr(self, "target_map_var") else self.cfg.targetMapPath
            target_map = validate_current_farm_map_path(Path(target_s), "目标正常 MAP") if target_s else None
            if maps:
                total_changes = 0
                self.log(f"本地批量替换：发现 {len(maps)} 个 .map 文件。")
                for local_map in maps:
                    _, _, changes = run_local_simulation(self.cfg, relations, local_map, target_map, extra, self.log)
                    total_changes += changes
                    self.log("------------------------------")
                self.log(f"本地批量替换完成：处理 {len(maps)} 个文件，替换 {total_changes} 条。")
            else:
                local_map, target_map = self.get_local_paths_for_sim()
                run_local_simulation(self.cfg, relations, local_map, target_map, extra, self.log)
        self.run_bg("本地批量替换", work)

    def task_test_remote(self):
        def work():
            self.save_settings_no_popup()
            RemoteClient(self.cfg, self.log).test()
        self.run_bg("测试连接", work)

    def task_download(self):
        def work():
            self.save_settings_no_popup()
            p = RemoteClient(self.cfg, self.log).download()
            # 下载后自动填成本机 MAP
            self.cfg.localMapPath = str(p)
            save_config(self.cfg)
            if hasattr(self, "local_map_var"):
                self.after(0, lambda: self.local_map_var.set(str(p)))
        self.run_bg("下载服务器 MAP", work)

    def _latest_update_file(self) -> Path:
        fixed = farm_runtime_path("update", self.cfg.remoteFile)
        if fixed.exists():
            return fixed
        maps = sorted(farm_runtime_dir("update").glob("*.map"), key=lambda p: p.stat().st_mtime, reverse=True)
        if not maps:
            raise RuntimeError("update 目录没有待上传 MAP。请先本地批量替换生成。")
        return maps[0]

    def task_upload_latest_update(self):
        relations = list(self.cloud_relations)
        def work():
            self.save_settings_no_popup()
            f = self._latest_update_file()
            client = RemoteClient(self.cfg, self.log)
            client.backup_remote(remote_backup_stem(relations))
            client.upload(f)
        self.run_bg("上传 update 目录 MAP", work)

    def task_one_click(self):
        relations = list(self.cloud_relations)
        def work():
            self.save_settings_no_popup()
            save_relations_for_scope("cloud", relations)
            client = RemoteClient(self.cfg, self.log)
            remote_backup = client.backup_remote(remote_backup_stem(relations))
            downloaded = client.download()
            backup = backup_original_name(downloaded)
            backup_meta = {
                "remote_file": self.cfg.remoteFile,
                "remote_dir": self.cfg.remoteDir,
                "remote_backup_file": remote_backup,
                "time": now_stamp(),
                "backup_file": str(backup),
                "pairs": [asdict(r) for r in relations],
            }
            farm_runtime_path("backup", "last_backup.json").write_text(json.dumps(backup_meta, ensure_ascii=False, indent=2), encoding="utf-8")
            self.log(f"一键仿真原始备份：{backup}")
            # 远程下载作为本机 MAP；如果目标 MAP 未指定，则同文件内部仿真
            self.cfg.localMapPath = str(downloaded)
            save_config(self.cfg)
            extra = self.extra_text.get("1.0", tk.END) if hasattr(self, "extra_text") else load_extra_rules_text()
            output, report, changes = run_local_simulation(self.cfg, relations, downloaded, Path(self.cfg.targetMapPath) if self.cfg.targetMapPath else None, extra, self.log)
            upload_file = farm_runtime_path("update", self.cfg.remoteFile)
            client.upload(upload_file if upload_file.exists() else output)
        self.run_bg("一键仿真", work)

    def task_restore_backup(self):
        def work():
            self.save_settings_no_popup()
            marker = farm_runtime_path("backup", "last_backup.json")
            latest: Optional[Path] = None
            remote_backup_file = ""
            if marker.exists():
                try:
                    meta = json.loads(marker.read_text(encoding="utf-8"))
                    backup_path = Path(str(meta.get("backup_file", "")))
                    if backup_path.exists():
                        latest = backup_path
                    remote_backup_file = str(meta.get("remote_backup_file", "")).strip()
                except Exception:
                    latest = None
                    remote_backup_file = ""
            if latest is None:
                backups = sorted(farm_runtime_dir("backup").glob("*.map"), key=lambda p: p.stat().st_mtime, reverse=True)
                if not backups:
                    raise RuntimeError("backup 目录没有备份 MAP，无法恢复。")
                latest = backups[0]
            self.log(f"使用最近备份恢复：{latest}")
            client = RemoteClient(self.cfg, self.log)
            client.upload(latest)
            if remote_backup_file:
                client.delete_remote_backup(remote_backup_file)
            else:
                self.log("未找到服务器备份记录，跳过删除服务器备份。")
        self.run_bg("取消仿真 / 恢复备份", work)

    def run_tool_action(self, kind: str):
        def work():
            self.save_settings_no_popup()
            c = RemoteClient(self.cfg, self.log)
            if kind == "winscp":
                c.open_winscp()
            elif kind == "flashfxp":
                c.open_flashfxp()
            elif kind == "omtg":
                c.open_omtg()
            elif kind == "ssh":
                c.open_ssh_terminal()
            elif kind == "sftp":
                c.open_sftp_terminal()
        self.run_bg(f"打开 {kind}", work)

    def import_mapping_gui(self):
        p = filedialog.askopenfilename(title="选择映射表", filetypes=[("映射表", "*.csv;*.xlsx;*.xls"), ("所有文件", "*.*")])
        if not p:
            return
        def work():
            count = import_mapping_table(Path(p))
            self.log(f"导入映射表成功：{count} 行")
            self.after(0, lambda: messagebox.showinfo("导入成功", f"已导入 {count} 行到 device_maps.csv"))
        self.run_bg("导入映射表", work)

    def show_mapping_count(self):
        try:
            entries = load_device_maps()
            fans = len(set(e.fan for e in entries))
            self.mapping_info_var.set(f"映射行数：{len(entries)}，风机数量：{fans}")
        except Exception as e:
            self.mapping_info_var.set(f"读取失败：{e}")

    def on_close(self):
        try:
            self.save_settings_no_popup()
            save_relations_for_scope("local", self.local_relations)
            save_relations_for_scope("cloud", self.cloud_relations)
            if hasattr(self, "extra_text"):
                save_extra_rules_text(self.extra_text.get("1.0", tk.END))
        except Exception:
            pass
        self.destroy()


# =====================================================================
# V4 固化规则兼容层：对齐前面 HTA 最后版的本地处理逻辑
# 重点：现场人员不需要导入规则。data/device_maps.csv 随包自带完整全场规则。
# 规则格式兼容 HTA：每行 = 风机名,地址1,地址2,...
# MAP 处理逻辑兼容 HTA：按 col2值|col3值|地址 建索引，默认 1|3|addr，只替换第 0 列。
# =====================================================================

LEGACY_FANS = [
    "F1-01FJ","F1-02FJ","F1-03FJ","F2-04FJ","F2-05FJ",
    "F3-06FJ","F3-07FJ","F3-08FJ","F4-09FJ","F4-10FJ",
    "F5-11FJ","F5-12FJ","F5-13FJ","F6-14FJ","F6-15FJ",
    "F7-16FJ","F7-17FJ","F7-18FJ","F8-19FJ","F8-20FJ",
]


def _create_legacy_default_device_maps() -> None:
    DEVICE_MAPS_PATH.parent.mkdir(parents=True, exist_ok=True)
    lines = ["# 每行格式：风机名,地址1,地址2,...  默认规则已内置，现场不用导入\n"]
    for i, fan in enumerate(LEGACY_FANS):
        b = i * 382
        vals = [str(x) for x in range(b, b + 381, 2)]
        vals.append(str(7688 + 2 * i))
        vals.append(str(7728 + 2 * i))
        lines.append(fan + "," + ",".join(vals) + "\n")
    DEVICE_MAPS_PATH.write_text("".join(lines), encoding="utf-8")


def ensure_default_files() -> None:  # type: ignore[override]
    ensure_dirs()
    if not DEVICE_MAPS_PATH.exists() or DEVICE_MAPS_PATH.stat().st_size < 1000:
        _create_legacy_default_device_maps()
    if not RELATIONS_PATH.exists():
        with RELATIONS_PATH.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(["enabled", "local_fan", "target_fan", "note"])
    if not EXTRA_RULES_PATH.exists():
        EXTRA_RULES_PATH.write_text(
            "# 可选额外行规则，正常现场不用改。\n"
            "# 格式：本机行=目标行，例如：900=1280\n"
            "# 也支持范围：900-901=1280-1281\n",
            encoding="utf-8",
        )


def load_legacy_device_maps() -> Dict[str, List[str]]:
    ensure_default_files()
    maps: Dict[str, List[str]] = {}
    for raw in DEVICE_MAPS_PATH.read_text(encoding="utf-8-sig", errors="ignore").splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        parts = [x.strip() for x in line.split(",")]
        if len(parts) < 2:
            continue
        fan = normalize_fan_name(parts[0])
        vals = [x for x in parts[1:] if x != ""]
        if fan and vals:
            maps[fan] = vals
    return maps


def _read_map_csv_lines(path: Path) -> Tuple[List[str], str]:
    data = path.read_bytes()
    for enc in ("gbk", "gb18030", "utf-8-sig", "utf-8", "latin1"):
        try:
            text = data.decode(enc)
            text = text.replace("\r\n", "\n").replace("\r", "\n")
            lines = text.split("\n")
            if lines and lines[-1] == "":
                lines = lines[:-1]
            return lines, enc
        except Exception:
            continue
    text = data.decode("gb18030", errors="replace").replace("\r\n", "\n").replace("\r", "\n")
    lines = text.split("\n")
    if lines and lines[-1] == "":
        lines = lines[:-1]
    return lines, "gb18030"


def _write_map_csv_lines(path: Path, lines: List[str], enc: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    text = "\r\n".join(lines)
    try:
        path.write_text(text, encoding=enc, newline="")
    except Exception:
        path.write_text(text, encoding="gb18030", newline="")


def _split_csv_simple(line: str) -> List[str]:
    return line.split(",")


def _legacy_build_index(lines: List[str]) -> Dict[str, int]:
    idx: Dict[str, int] = {}
    for i, line in enumerate(lines):
        c = _split_csv_simple(line)
        if len(c) < 5:
            continue
        k = f"{c[2]}|{c[3]}|{c[4]}"
        if k not in idx:
            idx[k] = i
    return idx


def _legacy_skip_line(line: str, skip_terms: str) -> bool:
    terms = [x for x in re.split(r"[,;|\s]+", skip_terms or "") if x]
    up = line.upper()
    for t in terms:
        if t.upper() in up:
            return True
    return False


def _legacy_process_one_file(cfg: Config, relations: List[Relation], in_path: Path, source_path: Optional[Path], out_path: Path, report_path: Path, log_func) -> int:
    lines, enc = _read_map_csv_lines(in_path)
    if source_path and source_path.exists() and source_path.resolve() != in_path.resolve():
        src_lines, _ = _read_map_csv_lines(source_path)
    else:
        src_lines = lines[:]
    dst_index = _legacy_build_index(lines)
    src_index = _legacy_build_index(src_lines)
    maps = load_legacy_device_maps()
    match_col2 = str(getattr(cfg, "matchCol2", "1") or "1")
    match_col3 = str(getattr(cfg, "matchCol3", "3") or "3")
    try:
        replace_col = int(getattr(cfg, "replaceColumnIndex", 0) or 0)
    except Exception:
        replace_col = 0
    skip_terms = str(getattr(cfg, "skipTerms", "IEMP") or "IEMP")
    rows = [["原始文件名", "替换关系", "本机风机", "目标风机", "本机地址", "目标地址", "替换前值", "替换后值", "是否变化", "对比", "规则来源"]]
    changed_count = 0
    enabled = [r for r in relations if r.enabled and r.local_fan and r.target_fan]
    if not enabled:
        raise RuntimeError("请先添加至少一组仿真关系。")
    for pair in enabled:
        lf = normalize_fan_name(pair.local_fan)
        sf = normalize_fan_name(pair.target_fan)
        if lf not in maps or sf not in maps:
            raise RuntimeError(f"地址映射不存在：{lf} / {sf}")
        lm, sm = maps[lf], maps[sf]
        if len(lm) != len(sm):
            raise RuntimeError(f"地址映射长度不一致：{lf} / {sf}")
        log_func(f"处理关系：{lf} 仿 {sf}，规则地址数量 {len(lm)}")
        for la, sa in zip(lm, sm):
            dk = f"{match_col2}|{match_col3}|{la}"
            sk = f"{match_col2}|{match_col3}|{sa}"
            if dk not in dst_index or sk not in src_index:
                continue
            di, si = dst_index[dk], src_index[sk]
            if _legacy_skip_line(lines[di], skip_terms) or _legacy_skip_line(src_lines[si], skip_terms):
                continue
            dst = _split_csv_simple(lines[di])
            src = _split_csv_simple(src_lines[si])
            if len(dst) <= replace_col or len(src) <= replace_col:
                continue
            oldv, newv = dst[replace_col], src[replace_col]
            changed = "是" if oldv != newv else "否"
            if oldv != newv:
                changed_count += 1
            dst[replace_col] = newv
            lines[di] = ",".join(dst)
            rows.append([in_path.name, f"{lf}仿{sf}", lf, sf, la, sa, oldv, newv, changed, f"{oldv} -> {newv}", "rules/" + get_current_wind_farm() + "/device_maps.csv"])
    extra_text = load_extra_rules_text()
    for local_line, target_line, raw_desc in parse_extra_rules(extra_text):
        li, ti = local_line - 1, target_line - 1
        if li < 0 or ti < 0 or li >= len(lines) or ti >= len(src_lines):
            continue
        dst = _split_csv_simple(lines[li])
        src = _split_csv_simple(src_lines[ti])
        if not dst or not src:
            continue
        oldv, newv = dst[0], src[0]
        changed = "是" if oldv != newv else "否"
        if oldv != newv:
            changed_count += 1
        dst[0] = newv
        lines[li] = ",".join(dst)
        rows.append([in_path.name, "额外项", "额外项", "额外项", str(local_line), str(target_line), oldv, newv, changed, f"{oldv} -> {newv}", raw_desc])
    _write_map_csv_lines(out_path, lines, enc)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    with report_path.open("w", encoding="gb18030", newline="") as f:
        w = csv.writer(f)
        w.writerows(rows)
    return changed_count


def _safe_output_part(value: str, fallback: str) -> str:
    text = str(value or "").strip()
    text = re.sub(r'[\\/:*?"<>|]+', "_", text)
    text = re.sub(r"\s+", "_", text)
    text = text.strip("._ ")
    return text or fallback


def _enabled_relation_list(relations: List[Relation]) -> List[Relation]:
    return [r for r in relations if r.enabled and r.local_fan and r.target_fan]


def _relation_output_label(rel: Relation) -> str:
    local = _safe_output_part(rel.local_fan, "local")
    target = _safe_output_part(rel.target_fan, "target")
    return f"{local}_from_{target}"


def _relations_output_label(relations: List[Relation]) -> str:
    enabled = _enabled_relation_list(relations)
    labels = [_relation_output_label(r) for r in enabled[:3]]
    if not labels:
        return "no_relation"
    if len(enabled) > 3:
        labels.append(f"{len(enabled)}_pairs")
    return "__".join(labels)


def _legacy_output_stem(file_name: str, relations: List[Relation], stamp: str) -> str:
    b = _safe_output_part(Path(file_name).stem, "map")
    farm = _safe_output_part(get_current_wind_farm(), "wind_farm")
    relation = _relations_output_label(relations)
    return f"{b}__{farm}__{relation}__{stamp}"


def _legacy_output_name(file_name: str, relations: List[Relation], stamp: Optional[str] = None) -> str:
    return f"{_legacy_output_stem(file_name, relations, stamp or now_stamp())}.map"


def _legacy_report_name(file_name: str, relations: List[Relation], stamp: Optional[str] = None) -> str:
    return f"{_legacy_output_stem(file_name, relations, stamp or now_stamp())}__report.csv"


def _legacy_summary_name(file_name: str, relations: List[Relation], stamp: Optional[str] = None) -> str:
    return f"{_legacy_output_stem(file_name, relations, stamp or now_stamp())}__summary.txt"


def run_local_simulation(cfg: Config, relations: List[Relation], local_map: Path, target_map: Optional[Path], extra_rules_text: str, log_func) -> Tuple[Path, Path, int]:  # type: ignore[override]
    if not local_map.exists():
        raise RuntimeError(f"本机 MAP 不存在：{local_map}")
    save_extra_rules_text(extra_rules_text)
    backup = backup_original_name(local_map) if local_map.name == cfg.remoteFile else backup_file(local_map, "local_before_sim")
    log_func(f"已备份本机 MAP：{backup}")
    stamp = now_stamp()
    enabled = _enabled_relation_list(relations)
    log_func(f"当前风场：{get_current_wind_farm()}")
    log_func(f"本次仿真关系：{len(enabled)} 组")
    for idx, rel in enumerate(enabled, 1):
        note = f"，说明：{rel.note}" if rel.note else ""
        log_func(f"  {idx}. {rel.local_fan} -> {rel.target_fan}{note}")
    output_map = farm_runtime_path("output_maps", _legacy_output_name(local_map.name, relations, stamp))
    report_csv = farm_runtime_path("reports", _legacy_report_name(local_map.name, relations, stamp))
    summary_txt = farm_runtime_path("reports", _legacy_summary_name(local_map.name, relations, stamp))
    changes = _legacy_process_one_file(cfg, relations, local_map, target_map, output_map, report_csv, log_func)
    update_map = farm_runtime_path("update", cfg.remoteFile)
    shutil.copy2(output_map, update_map)
    summary_lines = [
        f"风场: {get_current_wind_farm()}",
        f"输入MAP: {local_map}",
        f"目标MAP: {target_map if target_map else local_map}",
        f"输出MAP: {output_map}",
        f"待上传MAP: {update_map}",
        f"替换报告: {report_csv}",
        f"替换条数: {changes}",
        "仿真关系:",
    ]
    summary_lines.extend(
        f"{idx}. {rel.local_fan} -> {rel.target_fan}" + (f" | {rel.note}" if rel.note else "")
        for idx, rel in enumerate(enabled, 1)
    )
    summary_txt.write_text("\n".join(summary_lines) + "\n", encoding="utf-8")
    log_func(f"生成仿真摘要：{summary_txt}")
    log_func(f"生成仿真 MAP：{output_map}")
    log_func(f"同步待上传 MAP：{update_map}")
    log_func(f"生成替换报告：{report_csv}")
    log_func(f"替换条数：{changes}")
    return output_map, report_csv, changes


def _v4_import_mapping_gui(self):
    messagebox.showinfo("规则已内置", "V4 已随包内置全场规则库，现场人员正常不用导入。\n\n只有现场规则真的变更时，才需要由开发人员替换 data\\device_maps.csv。\n这不是日常操作。")


def _v4_show_mapping_count(self):
    try:
        maps = load_legacy_device_maps()
        total = sum(len(v) for v in maps.values())
        self.mapping_info_var.set(f"内置规则：{len(maps)} 台风机，{total} 个地址点，现场无需导入")
    except Exception as e:
        self.mapping_info_var.set(f"读取失败：{e}")

App.import_mapping_gui = _v4_import_mapping_gui  # type: ignore[name-defined]
App.show_mapping_count = _v4_show_mapping_count  # type: ignore[name-defined]


# =====================================================================
# V5 多风场规则库层
# 需求：首页选择风场；默认“红山嘴风电一场”；规则单独放 rules\风场名\；
# 一个风场一个子文件夹；导入规则只导入到当前风场，不影响其他风场。
# =====================================================================
DEFAULT_WIND_FARM = "红山嘴风电一场"
CURRENT_FARM_PATH = DIRS["data"] / "current_wind_farm.txt"


def _safe_farm_name(name: str) -> str:
    name = str(name or "").strip()
    name = re.sub(r'[\\/:*?"<>|]+', "_", name)
    return name or DEFAULT_WIND_FARM


def get_current_wind_farm() -> str:
    ensure_dirs()
    if CURRENT_FARM_PATH.exists():
        v = CURRENT_FARM_PATH.read_text(encoding="utf-8", errors="ignore").strip()
        if v:
            return _safe_farm_name(v)
    return DEFAULT_WIND_FARM


def set_current_wind_farm(name: str) -> None:
    ensure_dirs()
    name = _safe_farm_name(name)
    CURRENT_FARM_PATH.write_text(name + "\n", encoding="utf-8")
    ensure_wind_farm_profile(name)
    ensure_farm_runtime_dirs(name)


def get_farm_dir(name: Optional[str] = None) -> Path:
    return DIRS["rules"] / _safe_farm_name(name or get_current_wind_farm())


def get_current_farm_dir() -> Path:
    return get_farm_dir(get_current_wind_farm())


def validate_current_farm_map_path(path: Path, label: str) -> Path:
    path = path.expanduser()
    allowed = farm_runtime_dir("input_maps").resolve()
    try:
        resolved = path.resolve()
    except Exception:
        resolved = path
    try:
        resolved.relative_to(allowed)
    except ValueError:
        raise RuntimeError(f"{label} 不在当前风场目录：{allowed}")
    return resolved


def farm_device_maps_path(name: Optional[str] = None) -> Path:
    return get_farm_dir(name) / "device_maps.csv"


def farm_relations_path(name: Optional[str] = None) -> Path:
    return get_farm_dir(name) / "relations.csv"


def farm_relations_path_for_scope(scope: str, name: Optional[str] = None) -> Path:
    suffix = "cloud" if scope == "cloud" else "local"
    return get_farm_dir(name) / f"relations_{suffix}.csv"


def farm_extra_rules_path(name: Optional[str] = None) -> Path:
    return get_farm_dir(name) / "extra_rules.txt"


def ensure_wind_farm_profile(name: Optional[str] = None) -> None:
    name = _safe_farm_name(name or get_current_wind_farm())
    fd = get_farm_dir(name)
    fd.mkdir(parents=True, exist_ok=True)
    dm = fd / "device_maps.csv"
    rel = fd / "relations.csv"
    extra = fd / "extra_rules.txt"
    if not dm.exists() or dm.stat().st_size < 1000:
        try:
            _create_legacy_default_device_maps()
            if DEVICE_MAPS_PATH.exists():
                shutil.copy2(DEVICE_MAPS_PATH, dm)
        except Exception:
            lines = ["# 每行格式：风机名,地址1,地址2,...\n"]
            for i, fan in enumerate(LEGACY_FANS):
                b = i * 382
                vals = [str(x) for x in range(b, b + 381, 2)]
                vals.append(str(7688 + 2 * i))
                vals.append(str(7728 + 2 * i))
                lines.append(fan + "," + ",".join(vals) + "\n")
            dm.write_text("".join(lines), encoding="utf-8")
    if not rel.exists():
        with rel.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(["enabled", "local_fan", "target_fan", "note"])
    if not extra.exists():
        extra.write_text(
            "# 当前风场额外行规则。正常现场不用改。\n"
            "# 格式：本机行=目标行，例如：900=1280\n"
            "# 也支持范围：900-901=1280-1281\n",
            encoding="utf-8",
        )


def ensure_default_files() -> None:  # type: ignore[override]
    ensure_dirs()
    ensure_wind_farm_profile(DEFAULT_WIND_FARM)
    ensure_wind_farm_profile(get_current_wind_farm())
    ensure_farm_runtime_dirs(get_current_wind_farm())
    if not DEVICE_MAPS_PATH.exists() and farm_device_maps_path().exists():
        shutil.copy2(farm_device_maps_path(), DEVICE_MAPS_PATH)


def list_wind_farms() -> List[str]:  # type: ignore[override]
    ensure_dirs()
    ensure_wind_farm_profile(DEFAULT_WIND_FARM)
    farms = []
    for p in DIRS["rules"].iterdir():
        if p.is_dir() and (p / "device_maps.csv").exists():
            farms.append(p.name)
    if DEFAULT_WIND_FARM not in farms:
        farms.insert(0, DEFAULT_WIND_FARM)
    return sorted(set(farms), key=lambda x: (x != DEFAULT_WIND_FARM, x))


def load_relations() -> List[Relation]:  # type: ignore[override]
    ensure_wind_farm_profile(get_current_wind_farm())
    path = farm_relations_path()
    return load_relations_from_path(path)


def load_relations_from_path(path: Path) -> List[Relation]:
    rels: List[Relation] = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for r in reader:
            note = str(r.get("note", "")).strip()
            enabled = str(r.get("enabled", "1")).strip().lower() not in ("0", "false", "no", "否")
            if is_default_sample_relation(note):
                enabled = False
            rels.append(Relation(enabled, normalize_fan_name(r.get("local_fan", "")), normalize_fan_name(r.get("target_fan", "")), note))
    return rels


def save_relations(rels: List[Relation]) -> None:  # type: ignore[override]
    ensure_wind_farm_profile(get_current_wind_farm())
    path = farm_relations_path()
    save_relations_to_path(rels, path)


def save_relations_to_path(rels: List[Relation], path: Path) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["enabled", "local_fan", "target_fan", "note"])
        for r in rels:
            w.writerow(["1" if r.enabled else "0", r.local_fan, r.target_fan, r.note])


def sanitize_relations_for_current_farm(rels: List[Relation]) -> Tuple[List[Relation], bool]:
    try:
        fans = set(load_legacy_device_maps().keys())
    except Exception:
        return rels, False
    if not fans:
        return rels, False

    changed = False
    farm = get_current_wind_farm()
    for rel in rels:
        rel.local_fan = normalize_fan_name(rel.local_fan)
        rel.target_fan = normalize_fan_name(rel.target_fan)
        missing = [fan for fan in (rel.local_fan, rel.target_fan) if fan and fan not in fans]
        if rel.enabled and missing:
            rel.enabled = False
            rel.note = f"自动禁用：不属于当前风场 {farm}（{', '.join(missing)}）"
            changed = True
    return rels, changed


def ensure_scoped_relations_file(scope: str) -> Path:
    ensure_wind_farm_profile(get_current_wind_farm())
    path = farm_relations_path_for_scope(scope)
    if not path.exists():
        base = farm_relations_path()
        if base.exists():
            shutil.copy2(base, path)
        else:
            save_relations_to_path([], path)
    return path


def load_relations_for_scope(scope: str) -> List[Relation]:
    path = ensure_scoped_relations_file(scope)
    rels = load_relations_from_path(path)
    rels, changed = sanitize_relations_for_current_farm(rels)
    if changed:
        save_relations_to_path(rels, path)
    return rels


def save_relations_for_scope(scope: str, rels: List[Relation]) -> None:
    rels, _ = sanitize_relations_for_current_farm(rels)
    save_relations_to_path(rels, ensure_scoped_relations_file(scope))


def load_extra_rules_text() -> str:  # type: ignore[override]
    ensure_wind_farm_profile(get_current_wind_farm())
    return farm_extra_rules_path().read_text(encoding="utf-8", errors="ignore")


def save_extra_rules_text(text: str) -> None:  # type: ignore[override]
    ensure_wind_farm_profile(get_current_wind_farm())
    farm_extra_rules_path().write_text(text, encoding="utf-8")


def load_legacy_device_maps() -> Dict[str, List[str]]:  # type: ignore[override]
    ensure_wind_farm_profile(get_current_wind_farm())
    path = farm_device_maps_path()
    maps: Dict[str, List[str]] = {}
    for raw in path.read_text(encoding="utf-8-sig", errors="ignore").splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        parts = [x.strip() for x in line.split(",")]
        if len(parts) < 2:
            continue
        fan = normalize_fan_name(parts[0])
        vals = [x for x in parts[1:] if x != ""]
        if fan and vals:
            maps[fan] = vals
    return maps


def list_all_fans() -> List[str]:  # type: ignore[override]
    try:
        maps = load_legacy_device_maps()
        if maps:
            def fan_sort_key(name: str):
                return [int(x) if x.isdigit() else x for x in re.split(r"(\d+)", name)]
            return sorted(maps.keys(), key=fan_sort_key)
    except Exception:
        pass
    return LEGACY_FANS[:]


def wind_farm_summary() -> str:
    try:
        maps = load_legacy_device_maps()
        return f"当前：{get_current_wind_farm()} ｜ 风机 {len(maps)} 台 ｜ 地址点 {sum(len(v) for v in maps.values())} 个"
    except Exception as e:
        return f"当前：{get_current_wind_farm()} ｜ 规则读取失败：{e}"


def _normalize_import_rows_from_csv(path: Path) -> List[List[str]]:
    rows: List[List[str]] = []
    for raw in path.read_text(encoding="utf-8-sig", errors="ignore").splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        parts = [x.strip() for x in next(csv.reader([line]))]
        if len(parts) >= 2:
            rows.append(parts)
    return rows


def _normalize_import_rows_from_excel(path: Path) -> List[List[str]]:
    ext = path.suffix.lower()
    rows: List[List[str]] = []
    if ext in (".xlsx", ".xlsm"):
        import openpyxl  # type: ignore
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        for r in ws.iter_rows(values_only=True):
            vals = [str(x).strip() for x in r if x is not None and str(x).strip() != ""]
            if len(vals) >= 2:
                rows.append(vals)
    elif ext == ".xls":
        import xlrd  # type: ignore
        book = xlrd.open_workbook(str(path))
        sh = book.sheet_by_index(0)
        for rr in range(sh.nrows):
            vals = []
            for cc in range(sh.ncols):
                v = sh.cell_value(rr, cc)
                if v is None or str(v).strip() == "":
                    continue
                if isinstance(v, float) and v.is_integer():
                    vals.append(str(int(v)))
                else:
                    vals.append(str(v).strip())
            if len(vals) >= 2:
                rows.append(vals)
    return rows


def import_rules_to_current_wind_farm(src_path: Path) -> int:
    ensure_wind_farm_profile(get_current_wind_farm())
    ext = src_path.suffix.lower()
    if ext == ".csv":
        rows = _normalize_import_rows_from_csv(src_path)
    elif ext in (".xlsx", ".xlsm", ".xls"):
        rows = _normalize_import_rows_from_excel(src_path)
    else:
        raise RuntimeError("只支持 csv / xlsx / xls 规则表。")
    cleaned: List[List[str]] = []
    for r in rows:
        if not r:
            continue
        fan = normalize_fan_name(r[0])
        if not re.match(r"^F\d+-\d{2}FJ$", fan):
            continue
        addrs = [str(x).strip() for x in r[1:] if str(x).strip() != ""]
        if addrs:
            cleaned.append([fan] + addrs)
    if not cleaned:
        raise RuntimeError("没有识别到有效规则。推荐格式：每行 风机名,地址1,地址2,...，例如 F1-01FJ,0,2,4,6")
    dst = farm_device_maps_path()
    if dst.exists():
        bak = dst.with_name(f"device_maps_before_import_{now_stamp()}.csv")
        shutil.copy2(dst, bak)
    with dst.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow([f"# 风场：{get_current_wind_farm()}，每行格式：风机名,地址1,地址2,..."])
        for r in cleaned:
            w.writerow(r)
    return len(cleaned)


def _v5_switch_wind_farm(self):
    name = self.wind_farm_var.get().strip() if hasattr(self, "wind_farm_var") else get_current_wind_farm()
    if not name:
        return
    set_current_wind_farm(name)
    self.local_relations = load_relations_for_scope("local")
    self.cloud_relations = load_relations_for_scope("cloud")
    self.relations = self.cloud_relations if getattr(self, "relation_scope", "local") == "cloud" else self.local_relations
    self.refresh_relations_table()
    fans = list_all_fans()
    self.sync_fan_inputs_to_current_farm(fans)
    if hasattr(self, "extra_text"):
        self.extra_text.delete("1.0", tk.END)
        self.extra_text.insert("1.0", load_extra_rules_text())
    if hasattr(self, "wind_farm_info_var"):
        self.wind_farm_info_var.set(wind_farm_summary())
    self.log(f"已切换风场规则：{get_current_wind_farm()}，规则目录：{get_current_farm_dir()}")


def _v5_import_mapping_gui(self):
    p = filedialog.askopenfilename(
        initialdir=str(get_current_farm_dir()),
        title=f"给当前风场导入规则：{get_current_wind_farm()}",
        filetypes=[("规则表", "*.csv *.xlsx *.xls"), ("CSV", "*.csv"), ("Excel", "*.xlsx *.xls"), ("所有文件", "*.*")],
    )
    if not p:
        return
    try:
        count = import_rules_to_current_wind_farm(Path(p))
        messagebox.showinfo("导入成功", f"已导入 {count} 台风机规则到：\n{get_current_farm_dir()}")
        self.show_mapping_count()
        fans = list_all_fans()
        if hasattr(self, "local_fan_combo"):
            self.local_fan_combo.configure(values=fans)
        if hasattr(self, "target_fan_combo"):
            self.target_fan_combo.configure(values=fans)
    except Exception as e:
        messagebox.showerror("导入失败", str(e))


def _v5_show_mapping_count(self):
    try:
        maps = load_legacy_device_maps()
        total = sum(len(v) for v in maps.values())
        self.mapping_info_var.set(f"当前风场：{get_current_wind_farm()} ｜ 规则：{len(maps)} 台风机，{total} 个地址点")
        if hasattr(self, "wind_farm_info_var"):
            self.wind_farm_info_var.set(wind_farm_summary())
    except Exception as e:
        self.mapping_info_var.set(f"读取失败：{e}")


App.switch_wind_farm = _v5_switch_wind_farm  # type: ignore[name-defined]
App.import_mapping_gui = _v5_import_mapping_gui  # type: ignore[name-defined]
App.show_mapping_count = _v5_show_mapping_count  # type: ignore[name-defined]


# =====================================================================
# V6 风场规则模板层
# 目的：每个风场不只保存“风机=地址列表”，还保存 rule_profile.json。
# 这样后续其他风场即使 XLS 表结构、MAP 匹配列、固定匹配值、地址列、替换列不同，
# 也只改对应风场目录下的 rule_profile.json，不需要改程序源码。
# =====================================================================

def farm_rule_profile_path(name: Optional[str] = None) -> Path:
    return get_farm_dir(name) / "rule_profile.json"


def _default_rule_profile(name: Optional[str] = None) -> Dict[str, Any]:
    farm = _safe_farm_name(name or get_current_wind_farm())
    return {
        "profileVersion": 1,
        "windFarmName": farm,
        "description": "默认规则模板：兼容红山嘴风电一场和前面 HTA 最后版逻辑。正常只维护本文件和 device_maps.csv，不改源码。",
        "deviceMapFormat": {
            "type": "fan_address_list",
            "comment": "fan_address_list 表示每行：风机名,地址1,地址2,...。如果其他风场规则表结构不同，改 type 和下面列号即可。",
            "sheetNames": ["*"],
            "headerRows": 0,
            "fanColumn": 0,
            "addressStartColumn": 1,
            "addressColumn": 1,
            "excludeTailAddressCount": 0,
            "skipIfRowContains": ["IEMP"]
        },
        "mapParser": {
            "comment": "MAP 文件匹配规则。列号都是从 0 开始：0=第1列，1=第2列。默认 key = 第3列值 + 第4列值 + 第5列地址，固定匹配值为 1 和 3，只替换第1列。",
            "delimiter": ",",
            "keyColumnIndexes": [2, 3],
            "keyFixedValues": ["1", "3"],
            "addressColumnIndex": 4,
            "replaceColumnIndex": 0,
            "skipTerms": ["IEMP"]
        },
        "extraRules": {
            "enabled": True,
            "file": "extra_rules.txt",
            "replaceColumnIndex": 0,
            "lineNumberBase": 1
        }
    }


def _write_json(path: Path, obj: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")


def load_rule_profile(name: Optional[str] = None) -> Dict[str, Any]:
    path = farm_rule_profile_path(name)
    if not path.exists():
        _write_json(path, _default_rule_profile(name))
    try:
        data = json.loads(path.read_text(encoding="utf-8", errors="ignore"))
        if not isinstance(data, dict):
            raise ValueError("rule_profile.json 顶层必须是对象")
    except Exception:
        bak = path.with_name(f"rule_profile_broken_{now_stamp()}.json")
        try:
            shutil.copy2(path, bak)
        except Exception:
            pass
        data = _default_rule_profile(name)
        _write_json(path, data)
    base = _default_rule_profile(name)
    for k, v in base.items():
        if k not in data:
            data[k] = v
    for sec in ("deviceMapFormat", "mapParser", "extraRules"):
        if not isinstance(data.get(sec), dict):
            data[sec] = base[sec]
        else:
            for k, v in base[sec].items():
                data[sec].setdefault(k, v)
    return data


def ensure_wind_farm_profile(name: Optional[str] = None) -> None:  # type: ignore[override]
    name = _safe_farm_name(name or get_current_wind_farm())
    fd = get_farm_dir(name)
    fd.mkdir(parents=True, exist_ok=True)
    dm = fd / "device_maps.csv"
    rel = fd / "relations.csv"
    extra = fd / "extra_rules.txt"
    prof = fd / "rule_profile.json"
    if not prof.exists():
        _write_json(prof, _default_rule_profile(name))
    if not dm.exists() or dm.stat().st_size < 10:
        lines = ["# 每行格式默认：风机名,地址1,地址2,...；真正解析方式由 rule_profile.json 控制\n"]
        for i, fan in enumerate(LEGACY_FANS):
            b = i * 382
            vals = [str(x) for x in range(b, b + 381, 2)]
            vals.append(str(7688 + 2 * i))
            vals.append(str(7728 + 2 * i))
            lines.append(fan + "," + ",".join(vals) + "\n")
        dm.write_text("".join(lines), encoding="utf-8")
    if not rel.exists():
        with rel.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(["enabled", "local_fan", "target_fan", "note"])
    if not extra.exists():
        extra.write_text(
            "# 当前风场额外行规则。正常现场不用改。\n"
            "# 格式：本机行=目标行，例如：900=1280\n"
            "# 也支持范围：900-901=1280-1281\n",
            encoding="utf-8",
        )


def ensure_default_files() -> None:  # type: ignore[override]
    ensure_dirs()
    ensure_wind_farm_profile(DEFAULT_WIND_FARM)
    ensure_wind_farm_profile(get_current_wind_farm())
    ensure_farm_runtime_dirs(get_current_wind_farm())
    if not DEVICE_MAPS_PATH.exists() and farm_device_maps_path().exists():
        shutil.copy2(farm_device_maps_path(), DEVICE_MAPS_PATH)


def _as_int(v: Any, default: int) -> int:
    try:
        return int(v)
    except Exception:
        return default


def _as_list(v: Any) -> List[Any]:
    if isinstance(v, list):
        return v
    if v is None:
        return []
    return [v]


def _rule_skip_row(cells: List[str], profile: Dict[str, Any]) -> bool:
    terms = [str(x).upper() for x in _as_list(profile.get("deviceMapFormat", {}).get("skipIfRowContains", [])) if str(x).strip()]
    if not terms:
        return False
    joined = "|".join(cells).upper()
    return any(t in joined for t in terms)


def _parse_device_maps_rows(rows_by_sheet: Dict[str, List[List[str]]], profile: Dict[str, Any]) -> Dict[str, List[str]]:
    fmt = profile.get("deviceMapFormat", {})
    typ = str(fmt.get("type", "fan_address_list")).strip().lower()
    header_rows = max(0, _as_int(fmt.get("headerRows", 0), 0))
    fan_col = _as_int(fmt.get("fanColumn", 0), 0)
    addr_start = _as_int(fmt.get("addressStartColumn", 1), 1)
    addr_col = _as_int(fmt.get("addressColumn", 1), 1)
    sheet_names = [str(x) for x in _as_list(fmt.get("sheetNames", ["*"])) if str(x).strip()]
    exclude_tail = max(0, _as_int(fmt.get("excludeTailAddressCount", 0), 0))

    def sheet_allowed(sn: str) -> bool:
        if not sheet_names or "*" in sheet_names:
            return True
        return sn in sheet_names

    maps: Dict[str, List[str]] = {}
    for sheet, rows in rows_by_sheet.items():
        if not sheet_allowed(sheet):
            continue
        for row in rows[header_rows:]:
            cells = [str(x).strip() for x in row]
            if not cells or _rule_skip_row(cells, profile):
                continue
            if typ == "fan_address_list":
                if len(cells) <= fan_col:
                    continue
                fan = normalize_fan_name(cells[fan_col])
                if not fan:
                    continue
                addrs = [x for x in cells[addr_start:] if x != ""]
                if addrs:
                    maps.setdefault(fan, []).extend(addrs)
            elif typ == "fan_address_rows":
                if len(cells) <= max(fan_col, addr_col):
                    continue
                fan = normalize_fan_name(cells[fan_col])
                addr = cells[addr_col]
                if fan and addr:
                    maps.setdefault(fan, []).append(addr)
            elif typ == "sheet_as_fan":
                fan = normalize_fan_name(sheet)
                if len(cells) <= addr_col:
                    continue
                addr = cells[addr_col]
                if fan and addr:
                    maps.setdefault(fan, []).append(addr)
            else:
                raise RuntimeError(f"rule_profile.json 中 deviceMapFormat.type 不支持：{typ}")
    if exclude_tail:
        maps = {k: v[:-exclude_tail] if len(v) > exclude_tail else [] for k, v in maps.items()}
    return {k: v for k, v in maps.items() if v}


def load_legacy_device_maps() -> Dict[str, List[str]]:  # type: ignore[override]
    ensure_wind_farm_profile(get_current_wind_farm())
    path = farm_device_maps_path()
    profile = load_rule_profile(get_current_wind_farm())
    rows: List[List[str]] = []
    for raw in path.read_text(encoding="utf-8-sig", errors="ignore").splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        try:
            parts = [str(x).strip() for x in next(csv.reader([line]))]
        except Exception:
            parts = [x.strip() for x in line.split(",")]
        if parts:
            rows.append(parts)
    return _parse_device_maps_rows({"device_maps.csv": rows}, profile)


def _read_excel_sheets(path: Path) -> Dict[str, List[List[str]]]:
    ext = path.suffix.lower()
    out: Dict[str, List[List[str]]] = {}
    if ext in (".xlsx", ".xlsm"):
        import openpyxl  # type: ignore
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            rows: List[List[str]] = []
            for r in ws.iter_rows(values_only=True):
                vals = []
                for x in r:
                    if x is None:
                        vals.append("")
                    elif isinstance(x, float) and x.is_integer():
                        vals.append(str(int(x)))
                    else:
                        vals.append(str(x).strip())
                rows.append(vals)
            out[ws.title] = rows
    elif ext == ".xls":
        import xlrd  # type: ignore
        book = xlrd.open_workbook(str(path))
        for sh in book.sheets():
            rows = []
            for rr in range(sh.nrows):
                vals = []
                for cc in range(sh.ncols):
                    v = sh.cell_value(rr, cc)
                    if v is None:
                        vals.append("")
                    elif isinstance(v, float) and v.is_integer():
                        vals.append(str(int(v)))
                    else:
                        vals.append(str(v).strip())
                rows.append(vals)
            out[sh.name] = rows
    else:
        raise RuntimeError("只支持 xlsx / xlsm / xls。")
    return out


def _read_csv_sheet(path: Path) -> Dict[str, List[List[str]]]:
    rows: List[List[str]] = []
    data = path.read_bytes()
    text = None
    for enc in ("utf-8-sig", "gb18030", "gbk", "utf-8"):
        try:
            text = data.decode(enc)
            break
        except Exception:
            pass
    if text is None:
        text = data.decode("gb18030", errors="replace")
    for r in csv.reader(text.splitlines()):
        rows.append([str(x).strip() for x in r])
    return {path.name: rows}


def import_rules_to_current_wind_farm(src_path: Path) -> int:  # type: ignore[override]
    ensure_wind_farm_profile(get_current_wind_farm())
    profile = load_rule_profile(get_current_wind_farm())
    ext = src_path.suffix.lower()
    if ext == ".csv":
        rows_by_sheet = _read_csv_sheet(src_path)
    elif ext in (".xlsx", ".xlsm", ".xls"):
        rows_by_sheet = _read_excel_sheets(src_path)
    else:
        raise RuntimeError("只支持 csv / xlsx / xls 规则表。")
    maps = _parse_device_maps_rows(rows_by_sheet, profile)
    if not maps:
        raise RuntimeError("没有按当前风场 rule_profile.json 识别到有效规则。请先检查该风场的 rule_profile.json，而不是改程序源码。")
    dst = farm_device_maps_path()
    if dst.exists():
        bak = dst.with_name(f"device_maps_before_import_{now_stamp()}.csv")
        shutil.copy2(dst, bak)
    with dst.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow([f"# 风场：{get_current_wind_farm()}；本文件由导入生成；原始表结构解释见 rule_profile.json"])
        for fan in sorted(maps.keys()):
            w.writerow([fan] + maps[fan])
    return len(maps)


def _split_line_by_profile(line: str, profile: Dict[str, Any]) -> List[str]:
    delim = str(profile.get("mapParser", {}).get("delimiter", ",") or ",")
    if delim.lower() == "csv" or delim == ",":
        try:
            return [str(x) for x in next(csv.reader([line]))]
        except Exception:
            return line.split(",")
    if delim == "tab" or delim == "\\t":
        return line.split("\t")
    return line.split(delim)


def _join_line_by_profile(cells: List[str], profile: Dict[str, Any]) -> str:
    delim = str(profile.get("mapParser", {}).get("delimiter", ",") or ",")
    if delim.lower() == "csv" or delim == ",":
        return ",".join(cells)
    if delim == "tab" or delim == "\\t":
        return "\t".join(cells)
    return delim.join(cells)


def _profile_skip_line(line: str, profile: Dict[str, Any]) -> bool:
    terms = [str(x).upper() for x in _as_list(profile.get("mapParser", {}).get("skipTerms", [])) if str(x).strip()]
    up = line.upper()
    return any(t in up for t in terms)


def _build_map_index_by_profile(lines: List[str], profile: Dict[str, Any]) -> Dict[str, int]:
    mp = profile.get("mapParser", {})
    key_cols = [_as_int(x, -1) for x in _as_list(mp.get("keyColumnIndexes", [2, 3]))]
    addr_col = _as_int(mp.get("addressColumnIndex", 4), 4)
    idx: Dict[str, int] = {}
    for i, line in enumerate(lines):
        if _profile_skip_line(line, profile):
            continue
        c = _split_line_by_profile(line, profile)
        need = [addr_col] + key_cols
        if any(x < 0 or len(c) <= x for x in need):
            continue
        key_vals = [c[x] for x in key_cols]
        key = "|".join(key_vals + [c[addr_col]])
        if key not in idx:
            idx[key] = i
    return idx


def _key_for_address(addr: str, profile: Dict[str, Any]) -> str:
    mp = profile.get("mapParser", {})
    fixed = [str(x) for x in _as_list(mp.get("keyFixedValues", ["1", "3"]))]
    return "|".join(fixed + [str(addr)])


def _v6_process_one_file(cfg: Config, relations: List[Relation], in_path: Path, source_path: Optional[Path], out_path: Path, report_path: Path, log_func) -> int:
    profile = load_rule_profile(get_current_wind_farm())
    lines, enc = _read_map_csv_lines(in_path)
    if source_path and source_path.exists() and source_path.resolve() != in_path.resolve():
        src_lines, _ = _read_map_csv_lines(source_path)
    else:
        src_lines = lines[:]
    dst_index = _build_map_index_by_profile(lines, profile)
    src_index = _build_map_index_by_profile(src_lines, profile)
    maps = load_legacy_device_maps()
    replace_col = _as_int(profile.get("mapParser", {}).get("replaceColumnIndex", 0), 0)
    rows = [["原始文件名", "风场", "替换关系", "本机风机", "目标风机", "本机地址", "目标地址", "替换前值", "替换后值", "是否变化", "对比", "规则来源", "规则模板"]]
    changed_count = 0
    enabled = [r for r in relations if r.enabled and r.local_fan and r.target_fan]
    if not enabled:
        raise RuntimeError("请先添加至少一组仿真关系。")
    for pair in enabled:
        lf = normalize_fan_name(pair.local_fan)
        sf = normalize_fan_name(pair.target_fan)
        if lf not in maps or sf not in maps:
            raise RuntimeError(f"当前风场规则中地址映射不存在：{lf} / {sf}")
        lm, sm = maps[lf], maps[sf]
        if len(lm) != len(sm):
            raise RuntimeError(f"当前风场规则地址数量不一致：{lf}={len(lm)} / {sf}={len(sm)}")
        log_func(f"处理关系：{lf} 仿 {sf}，规则地址数量 {len(lm)}，风场：{get_current_wind_farm()}")
        for la, sa in zip(lm, sm):
            dk = _key_for_address(la, profile)
            sk = _key_for_address(sa, profile)
            if dk not in dst_index or sk not in src_index:
                continue
            di, si = dst_index[dk], src_index[sk]
            if _profile_skip_line(lines[di], profile) or _profile_skip_line(src_lines[si], profile):
                continue
            dst = _split_line_by_profile(lines[di], profile)
            src = _split_line_by_profile(src_lines[si], profile)
            if len(dst) <= replace_col or len(src) <= replace_col:
                continue
            oldv, newv = dst[replace_col], src[replace_col]
            changed = "是" if oldv != newv else "否"
            if oldv != newv:
                changed_count += 1
            dst[replace_col] = newv
            lines[di] = _join_line_by_profile(dst, profile)
            rows.append([in_path.name, get_current_wind_farm(), f"{lf}仿{sf}", lf, sf, la, sa, oldv, newv, changed, f"{oldv} -> {newv}", f"rules/{get_current_wind_farm()}/device_maps.csv", "rule_profile.json"])
    extra_text = load_extra_rules_text()
    extra_replace_col = _as_int(profile.get("extraRules", {}).get("replaceColumnIndex", replace_col), replace_col)
    for local_line, target_line, raw_desc in parse_extra_rules(extra_text):
        li, ti = local_line - 1, target_line - 1
        if li < 0 or ti < 0 or li >= len(lines) or ti >= len(src_lines):
            continue
        dst = _split_line_by_profile(lines[li], profile)
        src = _split_line_by_profile(src_lines[ti], profile)
        if len(dst) <= extra_replace_col or len(src) <= extra_replace_col:
            continue
        oldv, newv = dst[extra_replace_col], src[extra_replace_col]
        changed = "是" if oldv != newv else "否"
        if oldv != newv:
            changed_count += 1
        dst[extra_replace_col] = newv
        lines[li] = _join_line_by_profile(dst, profile)
        rows.append([in_path.name, get_current_wind_farm(), "额外项", "额外项", "额外项", str(local_line), str(target_line), oldv, newv, changed, f"{oldv} -> {newv}", raw_desc, "extra_rules.txt"])
    _write_map_csv_lines(out_path, lines, enc)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    with report_path.open("w", encoding="gb18030", newline="") as f:
        w = csv.writer(f)
        w.writerows(rows)
    return changed_count


_legacy_process_one_file = _v6_process_one_file  # type: ignore[assignment]


def wind_farm_summary() -> str:  # type: ignore[override]
    try:
        maps = load_legacy_device_maps()
        profile = load_rule_profile(get_current_wind_farm())
        mp = profile.get("mapParser", {})
        fmt = profile.get("deviceMapFormat", {})
        return f"当前：{get_current_wind_farm()} ｜ 风机 {len(maps)} 台 ｜ 地址点 {sum(len(v) for v in maps.values())} 个 ｜ 规则格式 {fmt.get('type')} ｜ MAP匹配值 {mp.get('keyFixedValues')}"
    except Exception as e:
        return f"当前：{get_current_wind_farm()} ｜ 规则读取失败：{e}"


def _v6_switch_wind_farm(self):
    name = self.wind_farm_var.get().strip() if hasattr(self, "wind_farm_var") else get_current_wind_farm()
    if not name:
        return
    set_current_wind_farm(name)
    self.local_relations = load_relations_for_scope("local")
    self.cloud_relations = load_relations_for_scope("cloud")
    self.relations = self.cloud_relations if getattr(self, "relation_scope", "local") == "cloud" else self.local_relations
    self.refresh_relations_table()
    fans = list_all_fans()
    self.sync_fan_inputs_to_current_farm(fans)
    if hasattr(self, "extra_text"):
        self.extra_text.delete("1.0", tk.END)
        self.extra_text.insert("1.0", load_extra_rules_text())
    if hasattr(self, "wind_farm_info_var"):
        self.wind_farm_info_var.set(wind_farm_summary())
    self.log(f"已切换风场规则：{get_current_wind_farm()}，规则目录：{get_current_farm_dir()}")
    self.log(f"规则模板：{farm_rule_profile_path()}")


def _v6_import_mapping_gui(self):
    p = filedialog.askopenfilename(
        initialdir=str(get_current_farm_dir()),
        title=f"按当前风场 rule_profile.json 导入规则：{get_current_wind_farm()}",
        filetypes=[("规则表", "*.csv *.xlsx *.xls"), ("CSV", "*.csv"), ("Excel", "*.xlsx *.xls"), ("所有文件", "*.*")],
    )
    if not p:
        return
    try:
        count = import_rules_to_current_wind_farm(Path(p))
        messagebox.showinfo("导入成功", f"已按当前风场 rule_profile.json 导入 {count} 台风机规则到：\n{get_current_farm_dir()}\n\n如果其他风场表结构不同，请先复制一个风场目录，再改它自己的 rule_profile.json。")
        self.show_mapping_count()
        fans = list_all_fans()
        if hasattr(self, "local_fan_combo"):
            self.local_fan_combo.configure(values=fans)
        if hasattr(self, "target_fan_combo"):
            self.target_fan_combo.configure(values=fans)
    except Exception as e:
        messagebox.showerror("导入失败", str(e))


def _v6_show_mapping_count(self):
    try:
        maps = load_legacy_device_maps()
        total = sum(len(v) for v in maps.values())
        profile = load_rule_profile(get_current_wind_farm())
        self.mapping_info_var.set(f"当前风场：{get_current_wind_farm()} ｜ 规则：{len(maps)} 台风机，{total} 个地址点 ｜ 模板：{profile.get('deviceMapFormat',{}).get('type')}")
        if hasattr(self, "wind_farm_info_var"):
            self.wind_farm_info_var.set(wind_farm_summary())
    except Exception as e:
        self.mapping_info_var.set(f"读取失败：{e}")


App.switch_wind_farm = _v6_switch_wind_farm  # type: ignore[name-defined]
App.import_mapping_gui = _v6_import_mapping_gui  # type: ignore[name-defined]
App.show_mapping_count = _v6_show_mapping_count  # type: ignore[name-defined]

def main():
    ensure_default_files()
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
